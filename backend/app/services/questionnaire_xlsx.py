"""The .xlsx pro-forma a designer builds their own questionnaire in, and the parser that reads it
back.

The loop this serves, in the owner's words: *designers want to make their OWN questionnaires*. They
download a pro-forma, type their questions into it in Excel, upload it, and then record answers
against it in the platform. The spreadsheet they upload MAY ALREADY CONTAIN ANSWERS — a designer who
ran the interview on paper and typed it up — and it may equally be blank. Both have to work, and
they have to produce the same shape of data, so an answer that arrived on the spreadsheet is stored
in exactly the same table as one typed into the app.

Two directions, and they are inverses of each other:

``build_pro_forma``           -> a blank workbook: the working sheet, a Details sheet for the title,
                                 and a worked example on the instructions sheet.
``build_questionnaire_workbook`` -> an EXISTING questionnaire as the same workbook, carrying question
                                 IDs and one Answer/Notes column pair per recorded sitting.
``build_question_set_workbook`` -> the same questionnaire's QUESTIONS AND NOTHING ELSE — no answers,
                                 no respondents, no question IDs. The artefact a designer SENDS to
                                 another designer; see the block comment above it.
``parse_questionnaire_workbook`` -> any of the above, or something a designer built from scratch,
                                 read back into sections + questions + answers.

THERE ARE THEREFORE TWO DOWNLOADS OF ONE QUESTIONNAIRE AND THEY ARE NOT INTERCHANGEABLE. The full
workbook is a LOSSLESS copy of somebody's fieldwork and belongs to the designer who collected it; the
question set is the INSTRUMENT, which is the part that is meant to travel. Conflating them is how a
sharing feature becomes a data leak, so they are two functions with two names rather than one
function with a flag — a flag defaults, and the wrong default here hands over a stranger's
respondents.

WHAT "FORGIVING" MEANS HERE, precisely, because it is the whole difference between a feature a
designer uses and one they abandon after the first upload:

* **Column order is not fixed.** Columns are located by reading the header row, not by position, so
  a designer who dragged "Answer" to the front loses nothing.
* **Header spelling is not fixed.** Each column accepts a family of names (see ``_COLUMN_ALIASES``);
  matching ignores case, punctuation, non-breaking spaces and repeated whitespace, because "Section
  Code", "section_code" and "SECTION  CODE " are the same column to everyone except a computer.
* **The header does not have to be on row 1.** Designers put a title across the top. The header row
  is *found* by scanning for one that names a question column.
* **The sheet does not have to be the first one.** A sheet named like the working sheet wins;
  failing that, whichever sheet yields the most questions does.
* **Section codes are optional.** A section is identified by its title if no code was typed, and a
  code is derived from the title so it stays stable when a designer inserts a section above it.
* **A section header can be written once or repeated.** Blank section cells inherit the last section
  seen, which is how a person lays a table out; repeating it on every row also works.

AND THE RULE THAT OUTRANKS ALL OF THAT: a row this parser cannot read is REPORTED, never dropped.
Every skipped or assumed-about row comes back in ``ParsedQuestionnaire.problems`` with its sheet, its
Excel row number and a sentence saying what happened. A silent drop means a designer uploads forty
questions, sees thirty-eight, and has no way to find out which two went missing or why — and the
most likely reason for a genuinely unreadable cell is the one case openpyxl cannot help with: a
FORMULA whose cached value was never written (see ``_cell_text``).

WRITING SIDE: every value goes into a cell through :func:`app.services.xlsx_report._put`, imported
rather than restated. That module's docstring records three separate ways ordinary field text made
Excel open a download with "We found a problem with some content" — a lone surrogate from a phone
that halved an emoji, a note beginning "=" that openpyxl stored as a formula, and a sheet name
carrying an apostrophe — and ``tests/test_xlsx_report.py`` is what keeps those fixed. A private copy
of that guard here would be a fourth renderer of the same file format that drifts the first time one
of them is corrected; the repo already shares ``csv_response`` between two routers for exactly this
reason. The import is deliberate, and if it ever breaks it breaks loudly at import time.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

# See the module docstring: shared, not restated. `_put` is the single door every value in an .xlsx
# this repo produces goes through, and it is what stops Excel offering to repair the download.
from app.services.xlsx_report import XLSX_MIME, _put, _sanitise

__all__ = [
    "PRO_FORMA_FILENAME",
    "QUESTION_SET_CONTENTS",
    "XLSX_MIME",
    "ParseProblem",
    "ParsedQuestion",
    "ParsedQuestionnaire",
    "ParsedSection",
    "QuestionnaireXlsxError",
    "build_pro_forma",
    "build_question_set_workbook",
    "build_questionnaire_workbook",
    "derive_section_code",
    "download_filename",
    "parse_questionnaire_workbook",
    "question_set_filename",
]

PRO_FORMA_FILENAME = "questionnaire-pro-forma.xlsx"

# Sheet names. The parser prefers the working sheet by name and the writer emits these, but neither
# depends on them: a designer who renamed the tab is still read (see `_pick_sheet`).
SHEET_QUESTIONS = "Questionnaire"
SHEET_DETAILS = "Details"
SHEET_HELP = "How to fill this in"

_BRAND = "5B21B6"  # the same purple the relational report's Overview uses
_WHITE = "FFFFFFFF"

# Column widths, in characters. The question column is where the typing happens, so it gets room.
_WIDTHS = {
    "sectionCode": 14,
    "sectionTitle": 30,
    "questionId": 26,
    "prompt": 62,
    "help": 34,
    "required": 11,
    "answer": 46,
    "notes": 34,
}

# Ceilings. A questionnaire is a research instrument, not a data dump: these exist so a malformed or
# hostile workbook cannot turn one upload into an unbounded read. They are far above any real form —
# the global artisan questionnaire, the largest instrument in the repository, has a few hundred
# questions across seventeen sections.
MAX_QUESTIONS = 2000
MAX_SECTIONS = 200
MAX_ROWS_SCANNED = 20000
MAX_SHEETS_SCANNED = 12
# How far down a sheet to look for the header row before concluding there isn't one. Generous
# enough for a title block and a logo, small enough that a sheet of prose is not mistaken for a form.
MAX_HEADER_SCAN_ROWS = 15

MAX_PROMPT_CHARS = 2000
MAX_TITLE_CHARS = 220
MAX_CODE_CHARS = 24
MAX_ANSWER_CHARS = 8000

_DEFAULT_SECTION_TITLE = "General"
_UPLOAD_ENTRY_LABEL = "Answer"


class QuestionnaireXlsxError(ValueError):
    """The upload is not a workbook this can read at all.

    Distinct from a *problem*, which is one row of an otherwise-good file. This is the whole file:
    not a zip, not an .xlsx, password-protected, or carrying no recognisable question column
    anywhere. The message is written to be shown to the designer as-is.
    """


# --- Reading: header vocabulary -----------------------------------------------------------------

# Non-breaking and friends. A header pasted out of a Word table or a browser routinely carries one of
# these instead of a space, and the column then matches nothing while looking identical on screen.
_SPACEY = re.compile(r"[\s  -\u200b  　]+")
# Punctuation a person adds to a header without meaning anything by it: "Question:", "Answer *",
# "Section (code)". Stripped before matching so all of those land on the same column.
_PUNCT = re.compile(r"[^\w\s]+", re.UNICODE)


def _norm_header(value: Any) -> str:
    """A header cell reduced to what it MEANS: 'Section  Code:' and 'section_code' both -> 'section code'."""
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = _PUNCT.sub(" ", text.replace("_", " "))
    return _SPACEY.sub(" ", text).strip().lower()


# Every spelling of every column this accepts. Order matters only within a role; roles are tried in
# the order of this dict, so the more specific ones ("section title") are matched before the looser
# ones ("section") could swallow them.
_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "sectionCode": ("section code", "sectioncode", "section ref", "section id", "code", "s no section"),
    "sectionTitle": ("section title", "section name", "sectiontitle", "section heading", "section", "part", "theme"),
    "questionId": ("question id", "questionid", "id", "q id", "qid", "ref", "reference", "question ref"),
    "prompt": ("question", "questions", "prompt", "question text", "the question", "item", "q"),
    "help": ("help", "help text", "helptext", "guidance", "hint", "instruction", "instructions", "description"),
    "required": ("required", "mandatory", "is required", "compulsory", "must answer"),
}
# Answer and note columns are matched by PREFIX, not exact name, because an exported questionnaire
# emits one pair per recorded sitting: "Answer — Ramesh", "Notes — Ramesh". The text after the dash
# names the sitting and is what pairs the two columns back up.
_ANSWER_PREFIXES = ("answer", "response", "reply")
_NOTES_PREFIXES = ("notes", "note", "remarks", "remark", "comment", "comments", "observation")
_LABEL_SPLIT = re.compile(r"\s*(?:[-–—:/|]|\bfor\b|\bby\b)\s*", re.IGNORECASE)

_TRUEISH = {"y", "yes", "true", "t", "1", "required", "mandatory", "compulsory", "x", "✓", "✔", "☑"}
_FALSEISH = {"n", "no", "false", "f", "0", "optional", "not required", "", "-", "—", "na", "n a"}


@dataclass(frozen=True)
class ParseProblem:
    """One thing the parser could not do cleanly, in terms a designer can act on.

    ``row`` is the 1-based worksheet row exactly as Excel's row gutter shows it, so "row 34" means
    press Ctrl+G and type 34. ``severity`` is ``"error"`` when nothing was stored for that row and
    ``"warning"`` when it was stored but something had to be assumed.
    """

    sheet: str
    row: int | None
    severity: str
    reason: str
    value: str | None = None

    def payload(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "row": self.row,
            "severity": self.severity,
            "reason": self.reason,
            "value": self.value,
        }


@dataclass
class ParsedQuestion:
    prompt: str
    row: int
    questionId: str | None = None
    helpText: str | None = None
    isRequired: bool = False
    # Sitting label -> answer text / note, for answers that arrived already filled in on the sheet.
    answers: dict[str, str] = field(default_factory=dict)
    answerNotes: dict[str, str] = field(default_factory=dict)


@dataclass
class ParsedSection:
    code: str
    title: str
    questions: list[ParsedQuestion] = field(default_factory=list)


@dataclass
class ParsedQuestionnaire:
    sections: list[ParsedSection] = field(default_factory=list)
    problems: list[ParseProblem] = field(default_factory=list)
    title: str | None = None
    description: str | None = None
    # Carried on the Details sheet of an exported questionnaire. Its presence is how a re-upload
    # knows it is an EDIT of a known form rather than a new one — though the endpoint takes the id
    # from the URL and only uses this to catch a designer uploading the wrong file.
    questionnaireId: str | None = None
    sheet: str | None = None
    # The sittings named by the answer columns, in the order they appeared.
    entryLabels: list[str] = field(default_factory=list)

    @property
    def questionCount(self) -> int:
        return sum(len(section.questions) for section in self.sections)

    @property
    def hasAnswers(self) -> bool:
        return any(q.answers for section in self.sections for q in section.questions)

    def payload(self) -> dict[str, Any]:
        return {
            "title": self.title,
            "description": self.description,
            "sheet": self.sheet,
            "sectionCount": len(self.sections),
            "questionCount": self.questionCount,
            "entryLabels": self.entryLabels,
            "problems": [p.payload() for p in self.problems],
        }


# --- Reading: cells -----------------------------------------------------------------------------


class _FormulaCell(str):
    """Marker for a cell that holds a formula whose value was never cached.

    openpyxl in ``data_only`` mode returns the value Excel last *calculated* and stored in the file.
    A workbook written by a script — LibreOffice headless, a generator, openpyxl itself — has
    formulas but no cached results, so that value is ``None`` and the cell reads as empty. Reporting
    "row 12 is blank" for a row that visibly contains ``=B12&" (2024)"`` on the designer's screen is
    the single most confusing thing this parser could say, so formula cells are detected on a second,
    non-evaluating pass and reported for what they are.
    """


def _cell_text(value: Any) -> str:
    """One cell as trimmed text. ``None``, whitespace and Excel's ``#N/A`` family all read as empty."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float) and value.is_integer():
        # Excel stores every number as a float, so a question numbered 3 arrives as 3.0 and a
        # question ID typed as a number would become "3.0" in the prompt.
        value = int(value)
    text = str(value).strip()
    if text.startswith("#") and text.endswith("!") and len(text) <= 10:
        return ""  # #REF!, #NAME?, #VALUE! — an error, not content
    if text in {"#N/A", "#NULL!"}:
        return ""
    return text


def _clip(text: str, limit: int) -> str:
    return text if len(text) <= limit else text[: limit - 1] + "…"


def _truthy(text: str) -> bool | None:
    """'Yes'/'x'/'✓' -> True, 'No'/'' -> False, anything else -> None (i.e. tell the designer)."""
    key = _norm_header(text)
    if key in _TRUEISH:
        return True
    if key in _FALSEISH:
        return False
    return None


def derive_section_code(title: str, taken: set[str]) -> str:
    """A stable, unique code for a section the designer titled but did not code.

    Derived from the TITLE rather than from the position. Positional codes ("S1", "S2") look tidier
    and are wrong: inserting a section at the top renumbers every section below it, so a re-upload
    matches nothing and the whole form churns. A title-derived code survives insertion.
    """
    base = re.sub(r"[^A-Za-z0-9]+", "_", unicodedata.normalize("NFKD", title)).strip("_").upper()
    base = base[:MAX_CODE_CHARS].strip("_") or "SECTION"
    candidate = base
    n = 2
    while candidate.lower() in taken:
        suffix = f"_{n}"
        candidate = base[: MAX_CODE_CHARS - len(suffix)] + suffix
        n += 1
    taken.add(candidate.lower())
    return candidate


# --- Reading: the header row --------------------------------------------------------------------


def _role_for(header: str) -> tuple[str, str | None] | None:
    """Map one normalised header to (role, sitting label), or None if it is not a column we know."""
    if not header:
        return None
    for role, aliases in _COLUMN_ALIASES.items():
        if header in aliases:
            return (role, None)
    for prefix in _ANSWER_PREFIXES:
        if header == prefix or header.startswith(prefix + " "):
            return ("answer", _label_after(header, prefix))
    for prefix in _NOTES_PREFIXES:
        if header == prefix or header.startswith(prefix + " "):
            return ("notes", _label_after(header, prefix))
    return None


def _label_after(header: str, prefix: str) -> str | None:
    """'answer ramesh devi' after prefix 'answer' -> 'Ramesh Devi'; a bare 'answer' -> None."""
    rest = header[len(prefix) :].strip()
    rest = _LABEL_SPLIT.sub(" ", rest).strip()
    return rest.title() if rest else None


def _scan_header(rows: list[tuple[int, list[str]]]) -> tuple[int, dict[int, tuple[str, str | None]]] | None:
    """Find the header row and what each of its columns means.

    A row qualifies only if it names a QUESTION column. Nothing else is sufficient: a sheet whose
    first row happens to read "Notes" is not a questionnaire, and treating it as one would produce a
    form full of empty prompts rather than an honest "I could not find the question column".
    """
    for row_number, values in rows[:MAX_HEADER_SCAN_ROWS]:
        mapping: dict[int, tuple[str, str | None]] = {}
        seen_roles: set[str] = set()
        for index, raw in enumerate(values):
            role = _role_for(_norm_header(raw))
            if role is None:
                continue
            # A repeated single-value column (two "Section Title"s) keeps the first; answer and note
            # columns legitimately repeat, one pair per sitting, so they are exempt.
            if role[0] not in ("answer", "notes") and role[0] in seen_roles:
                continue
            seen_roles.add(role[0])
            mapping[index] = role
        if "prompt" in seen_roles:
            return row_number, mapping
    return None


def _sheet_rows(ws: Worksheet, formulas: dict[tuple[int, int], str]) -> list[tuple[int, list[str]]]:
    """One sheet as (row number, [cell text, ...]), trailing blanks kept so column indices line up."""
    out: list[tuple[int, list[str]]] = []
    for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_number > MAX_ROWS_SCANNED:
            break
        cells: list[str] = []
        for column, value in enumerate(row, start=1):
            text = _cell_text(value)
            if not text:
                formula = formulas.get((row_number, column))
                if formula:
                    # Reads as empty here but is not empty on the designer's screen. Carried through
                    # as a marker so the row can be reported precisely instead of as "blank".
                    cells.append(_FormulaCell(formula))
                    continue
            cells.append(text)
        out.append((row_number, cells))
    return out


def _formula_map(data: bytes) -> dict[str, dict[tuple[int, int], str]]:
    """Every cell in the workbook that holds a formula, per sheet, from a non-evaluating load.

    A second parse of the same bytes. It buys the difference between "row 12 is blank" and "row 12
    holds a formula Excel has never calculated" — see ``_FormulaCell``. Cheap in the normal case,
    where the answer is an empty dict for every sheet.
    """
    found: dict[str, dict[tuple[int, int], str]] = {}
    try:
        wb = load_workbook(BytesIO(data), data_only=False, read_only=True)
    except (InvalidFileException, OSError, ValueError, KeyError, TypeError):
        # The evaluating load has already succeeded, so this one is expected to as well. If it does
        # not, the caller still gets a fully parsed questionnaire and merely loses the ability to
        # say "that cell is a formula" — a diagnostic nicety must never be the thing that fails an
        # upload the designer's file was fine for.
        return found
    try:
        for ws in wb.worksheets[:MAX_SHEETS_SCANNED]:
            per_sheet: dict[tuple[int, int], str] = {}
            for row_number, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_number > MAX_ROWS_SCANNED:
                    break
                for column, value in enumerate(row, start=1):
                    if isinstance(value, str) and value.startswith("="):
                        per_sheet[(row_number, column)] = value
            found[ws.title] = per_sheet
    finally:
        wb.close()
    return found


# --- Reading: the details sheet -----------------------------------------------------------------

_DETAIL_KEYS = {
    "title": ("questionnaire title", "title", "name", "questionnaire name", "form title"),
    "description": ("description", "purpose", "about", "notes", "summary"),
    "questionnaireId": ("questionnaire id", "id", "questionnaire ref", "form id"),
}


def _read_details(rows: list[tuple[int, list[str]]]) -> dict[str, str]:
    """Label/value pairs off the Details sheet. Tolerates the value being in any column to the right,
    which is what happens the moment a designer widens column A or inserts one."""
    found: dict[str, str] = {}
    for _row_number, values in rows[:60]:
        if not values:
            continue
        key = _norm_header(values[0])
        for name, aliases in _DETAIL_KEYS.items():
            if key in aliases and name not in found:
                value = next((v for v in values[1:] if v and not isinstance(v, _FormulaCell)), "")
                if value:
                    found[name] = value
    return found


# --- Reading: the public entry point ------------------------------------------------------------

# What the first few bytes of a file say it really is. Checked instead of the extension because the
# extension is the thing most likely to be wrong — a designer who renamed "report.xls" to
# "report.xlsx" to get past an upload filter has an .xls with an .xlsx name, and the message has to
# describe the file rather than the label somebody put on it.
_ZIP_MAGIC = b"PK\x03\x04"
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"  # .xls (BIFF8), and also an ENCRYPTED .xlsx


def _unreadable_file_message(data: bytes) -> str:
    """The sentence a designer is shown when the whole file cannot be opened."""
    head = data[:8]
    if head.startswith(_OLE2_MAGIC):
        # Both cases land here and the remedy differs, so both are named. Re-saving fixes the first
        # and removing the password fixes the second; neither is guessable from "invalid file".
        return (
            "That file is in the older .xls format, or it is an .xlsx with a password on it. Open "
            "it in Excel, remove any password, then use File > Save As and choose "
            "'Excel Workbook (.xlsx)'."
        )
    if not head.startswith(_ZIP_MAGIC):
        return (
            "That is not an Excel workbook — a .csv, a .numbers file or a PDF will not do. Fill in "
            "the .xlsx pro-forma, or use File > Save As and choose 'Excel Workbook (.xlsx)'."
        )
    return (
        "The workbook could not be opened — it may be password-protected or the upload may have "
        "been cut short. Open it in Excel, use File > Save As to save a fresh copy, and upload that."
    )


def parse_questionnaire_workbook(data: bytes, *, filename: str | None = None) -> ParsedQuestionnaire:
    """Read an uploaded workbook into sections + questions (+ any answers it already carried).

    Raises :class:`QuestionnaireXlsxError` only when the FILE is unusable. Anything wrong with a row
    of an otherwise-readable file comes back in ``result.problems`` and the rest of the file is
    still imported, because a designer with thirty-eight of forty questions and a list of the two
    that need fixing is far better off than one with an error page.
    """
    if not data:
        raise QuestionnaireXlsxError("The upload was empty. Attach the filled-in .xlsx pro-forma.")
    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError, KeyError, TypeError) as exc:
        # Which of those it was tells the designer nothing, so the message is chosen from the FILE'S
        # OWN MAGIC BYTES instead of from the exception class or the extension. openpyxl does not
        # raise its own InvalidFileException for a BytesIO — it has no filename to inspect — so an
        # old .xls arrives here as a bare zipfile.BadZipFile, and "the upload may have been cut
        # short" sends somebody to re-download a file that was never the problem.
        raise QuestionnaireXlsxError(_unreadable_file_message(data)) from exc

    try:
        formulas = _formula_map(data)
        sheets: dict[str, list[tuple[int, list[str]]]] = {}
        for ws in wb.worksheets[:MAX_SHEETS_SCANNED]:
            sheets[ws.title] = _sheet_rows(ws, formulas.get(ws.title, {}))
    finally:
        wb.close()

    if not sheets:
        raise QuestionnaireXlsxError("The workbook has no sheets in it.")

    result = ParsedQuestionnaire()

    # Details, if there is a sheet that looks like one. Missing is normal — a designer who built the
    # form from scratch has no Details sheet, and the title then comes from the filename.
    for name, rows in sheets.items():
        if _norm_header(name) in ("details", "about", "questionnaire details", "info", "information"):
            details = _read_details(rows)
            result.title = _clip(details["title"], MAX_TITLE_CHARS) if details.get("title") else None
            result.description = details.get("description") or None
            result.questionnaireId = details.get("questionnaireId") or None
            break

    picked = _pick_sheet(sheets, result.problems)
    if picked is None:
        raise QuestionnaireXlsxError(
            "No sheet in that workbook has a 'Question' column, so there was nothing to import. "
            "Download the pro-forma and type your questions under its headings, or add a header row "
            "with a column called 'Question'."
        )
    sheet_name, header_row, columns = picked
    result.sheet = sheet_name
    _read_questions(sheets[sheet_name], sheet_name, header_row, columns, result)

    if not result.title and filename:
        # Last resort, and a good one: designers name the file after the questionnaire.
        stem = re.sub(r"\.(xlsx|xlsm|xltx)$", "", filename, flags=re.IGNORECASE).strip()
        stem = re.sub(r"[_-]+", " ", stem).strip()
        if stem:
            result.title = _clip(stem, MAX_TITLE_CHARS)

    if not result.questionCount:
        raise QuestionnaireXlsxError(
            f"The '{sheet_name}' sheet has a Question column but no questions under it. Type at "
            "least one question and upload again."
        )
    return result


def _pick_sheet(
    sheets: dict[str, list[tuple[int, list[str]]]],
    problems: list[ParseProblem],
) -> tuple[str, int, dict[int, tuple[str, str | None]]] | None:
    """Which sheet holds the form. Name first, then whichever candidate has the most rows under its
    header — a designer whose working sheet is called "Sheet1" or "Final v3" is still read."""
    candidates: list[tuple[str, int, dict[int, tuple[str, str | None]]]] = []
    for name, rows in sheets.items():
        found = _scan_header(rows)
        if found is not None:
            candidates.append((name, found[0], found[1]))
    if not candidates:
        return None
    for candidate in candidates:
        name = candidate[0]
        if _norm_header(name) in (_norm_header(SHEET_QUESTIONS), "questions", "form", "sheet1"):
            if len(candidates) > 1:
                others = [c[0] for c in candidates if c[0] != name]
                problems.append(
                    ParseProblem(
                        sheet=name,
                        row=None,
                        severity="warning",
                        reason=(
                            "More than one sheet looked like a questionnaire; this one was used. "
                            f"Ignored: {', '.join(others)}."
                        ),
                    )
                )
            return candidate
    best = max(candidates, key=lambda c: len(sheets[c[0]]) - c[1])
    if len(candidates) > 1:
        others = [c[0] for c in candidates if c[0] != best[0]]
        problems.append(
            ParseProblem(
                sheet=best[0],
                row=None,
                severity="warning",
                reason=(
                    "No sheet was named 'Questionnaire', so the one with the most questions was "
                    f"used. Ignored: {', '.join(others)}."
                ),
            )
        )
    return best


def _read_questions(
    rows: list[tuple[int, list[str]]],
    sheet: str,
    header_row: int,
    columns: dict[int, tuple[str, str | None]],
    result: ParsedQuestionnaire,
) -> None:
    """Walk the data rows under the header, building sections as they are met."""
    sections: dict[str, ParsedSection] = {}
    order: list[str] = []
    codes_taken: set[str] = set()
    current: ParsedSection | None = None
    seen_ids: set[str] = set()
    entry_labels: list[str] = []
    # Counted as we go rather than read off `result`, whose sections are only assembled at the end —
    # reading the ceiling off an always-empty list is how this cap silently never fires.
    kept = 0

    def get(values: list[str], role: str, label: str | None = None) -> str:
        for index, (this_role, this_label) in columns.items():
            if this_role == role and this_label == label and index < len(values):
                value = values[index]
                return "" if isinstance(value, _FormulaCell) else value
        return ""

    def formula_in(values: list[str]) -> str | None:
        for index in sorted(columns):
            if index < len(values) and isinstance(values[index], _FormulaCell):
                return str(values[index])
        return None

    def section_for(code: str, title: str, row_number: int) -> ParsedSection:
        nonlocal current
        key = (code or derive_section_code(title, set())).lower()
        existing = sections.get(key)
        if existing is not None:
            if title and existing.title != title:
                result.problems.append(
                    ParseProblem(
                        sheet=sheet,
                        row=row_number,
                        severity="warning",
                        reason=(
                            f"Section '{existing.code}' is titled '{existing.title}' earlier in the "
                            f"sheet; the first title was kept."
                        ),
                        value=title,
                    )
                )
            current = existing
            return existing
        if len(sections) >= MAX_SECTIONS:
            result.problems.append(
                ParseProblem(
                    sheet=sheet,
                    row=row_number,
                    severity="error",
                    reason=f"More than {MAX_SECTIONS} sections; the rest of the sheet was ignored.",
                )
            )
            raise _StopReading
        final_code = code or derive_section_code(title, codes_taken)
        if code:
            codes_taken.add(code.lower())
        made = ParsedSection(code=final_code, title=title or final_code)
        sections[key] = made
        order.append(key)
        current = made
        return made

    for row_number, values in rows:
        if row_number <= header_row:
            continue
        prompt = get(values, "prompt")
        code = _clip(get(values, "sectionCode"), MAX_CODE_CHARS)
        title = _clip(get(values, "sectionTitle"), MAX_TITLE_CHARS)

        if not prompt and not code and not title:
            formula = formula_in(values)
            if formula:
                result.problems.append(
                    ParseProblem(
                        sheet=sheet,
                        row=row_number,
                        severity="error",
                        reason=(
                            "This row holds a formula whose result is not saved in the file, so "
                            "there was nothing to read. Open the workbook in Excel and save it "
                            "again, or replace the formula with the text it produces."
                        ),
                        value=_clip(formula, 120),
                    )
                )
            continue  # an ordinary blank row: layout, not data

        try:
            if not prompt:
                # A section header row: names a section and asks nothing. Common and correct.
                section_for(code, title, row_number)
                continue
            if code or title:
                section_for(code, title, row_number)
            if current is None:
                current = section_for("", _DEFAULT_SECTION_TITLE, row_number)
                result.problems.append(
                    ParseProblem(
                        sheet=sheet,
                        row=row_number,
                        severity="warning",
                        reason=(
                            "This question came before any section was named, so it was filed under "
                            f"'{_DEFAULT_SECTION_TITLE}'. Add a Section Title to move it."
                        ),
                        value=_clip(prompt, 120),
                    )
                )
        except _StopReading:
            break

        if kept >= MAX_QUESTIONS:
            result.problems.append(
                ParseProblem(
                    sheet=sheet,
                    row=row_number,
                    severity="error",
                    reason=f"More than {MAX_QUESTIONS} questions; the rest of the sheet was ignored.",
                )
            )
            break

        if len(prompt) > MAX_PROMPT_CHARS:
            result.problems.append(
                ParseProblem(
                    sheet=sheet,
                    row=row_number,
                    severity="warning",
                    reason=f"The question was longer than {MAX_PROMPT_CHARS} characters and was shortened.",
                    value=_clip(prompt, 120),
                )
            )
            prompt = _clip(prompt, MAX_PROMPT_CHARS)

        question_id = get(values, "questionId") or None
        if question_id and question_id in seen_ids:
            result.problems.append(
                ParseProblem(
                    sheet=sheet,
                    row=row_number,
                    severity="warning",
                    reason=(
                        f"Question ID '{question_id}' appears more than once. This row was imported "
                        "as a NEW question rather than overwriting the earlier one."
                    ),
                    value=_clip(prompt, 120),
                )
            )
            question_id = None
        if question_id:
            seen_ids.add(question_id)

        required_raw = get(values, "required")
        required = _truthy(required_raw)
        if required is None:
            result.problems.append(
                ParseProblem(
                    sheet=sheet,
                    row=row_number,
                    severity="warning",
                    reason=(
                        f"'{_clip(required_raw, 40)}' is not a yes/no value, so the question was "
                        "imported as optional. Use Yes or No."
                    ),
                    value=_clip(prompt, 120),
                )
            )
            required = False

        question = ParsedQuestion(
            prompt=prompt,
            row=row_number,
            questionId=question_id,
            helpText=get(values, "help") or None,
            isRequired=required,
        )
        for index, (role, label) in sorted(columns.items()):
            if role not in ("answer", "notes") or index >= len(values):
                continue
            raw = values[index]
            if isinstance(raw, _FormulaCell):
                result.problems.append(
                    ParseProblem(
                        sheet=sheet,
                        row=row_number,
                        severity="warning",
                        reason=(
                            "The answer on this row is a formula whose result is not saved in the "
                            "file, so the question was imported without it."
                        ),
                        value=_clip(prompt, 120),
                    )
                )
                continue
            if not raw:
                continue
            name = label or _UPLOAD_ENTRY_LABEL
            if name not in entry_labels:
                entry_labels.append(name)
            target = question.answers if role == "answer" else question.answerNotes
            target[name] = _clip(raw, MAX_ANSWER_CHARS)
        current.questions.append(question)
        kept += 1

    # Every section is kept, INCLUDING one that ended up with no questions: a designer who typed
    # their section headings first and their questions second would otherwise upload the file and
    # find the headings gone.
    result.sections = [sections[key] for key in order]
    result.entryLabels = entry_labels


class _StopReading(Exception):
    """Internal: a ceiling was hit and the rest of the sheet is deliberately not read."""


# --- Writing ------------------------------------------------------------------------------------


def _header(ws: Worksheet, row: int, labels: list[tuple[str, str]]) -> None:
    fill = PatternFill(fill_type="solid", start_color=f"FF{_BRAND}")
    for index, (key, label) in enumerate(labels, start=1):
        cell = _put(ws, row, index, label)
        cell.font = Font(bold=True, color=_WHITE)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        letter = ws.cell(row=row, column=index).column_letter
        ws.column_dimensions[letter].width = _WIDTHS.get(key, 24)
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = f"A{row + 1}"


_DETAILS_NOTE = (
    "Type your questionnaire's name and purpose above. Leave 'Questionnaire ID' exactly as "
    "you found it — it is how the app recognises this file as an edit of a questionnaire you "
    "have already uploaded rather than a brand new one."
)


def _details_sheet(
    ws: Worksheet,
    *,
    title: str,
    description: str,
    questionnaire_id: str,
    version: int | None,
    extra: list[tuple[str, str]] | None = None,
    note: str = _DETAILS_NOTE,
) -> None:
    """The Details sheet: four label/value rows, optionally some more, then one italic note.

    ``extra`` LABELS MUST NOT COLLIDE WITH ``_DETAIL_KEYS`` above, and that is a real constraint
    rather than a stylistic one — this sheet is parsed back by :func:`_read_details`, which matches
    the label in column A against a family of aliases. A row labelled "Notes" or "Summary" would be
    read back as the questionnaire's DESCRIPTION, so a provenance line added here for the reader's
    benefit would silently overwrite the description of the questionnaire the file is imported into.
    ``tests/test_questionnaire_interchange.py`` pins that the question set's own extra rows survive
    the round trip without doing this.

    The note row is placed BELOW whatever rows exist rather than at a fixed row 8, so adding a pair
    cannot land the paragraph on top of it. With the four base pairs it still lands on row 8, exactly
    where it always did.
    """
    ws.sheet_properties.tabColor = f"FF{_BRAND}"
    heading = _put(ws, 1, 1, "Questionnaire details")
    heading.font = Font(bold=True, size=14, color=f"FF{_BRAND}")
    pairs = [
        ("Questionnaire title", title),
        ("Description", description),
        ("Questionnaire ID", questionnaire_id),
        ("Version", str(version) if version is not None else ""),
        *(extra or []),
    ]
    for offset, (label, value) in enumerate(pairs):
        row = 3 + offset
        cell = _put(ws, row, 1, label)
        cell.font = Font(bold=True)
        value_cell = _put(ws, row, 2, value)
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")
    note_row = 3 + len(pairs) + 1
    note_cell = _put(ws, note_row, 1, note)
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    note_cell.font = Font(italic=True, color="FF6B7280")
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row + 2, end_column=4)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 62


_HELP_LINES = [
    ("h", "How to fill this in"),
    ("p", "Type your questions on the 'Questionnaire' sheet, one per row, then upload the file."),
    ("b", "Sections — put the section's name in 'Section Title'. Fill it in on the first question of the section and leave it blank on the rest, or repeat it on every row; both work. Leave 'Section Code' blank and the app will create one for you."),
    ("b", "Questions — 'Question' is the only column you must fill in. Every other column is optional."),
    ("b", "Required — type Yes or No. Anything else is read as No and reported back to you."),
    ("b", "Answers — you can leave the 'Answer' column empty and record answers in the app, or type answers you already collected on paper straight into it. Both work, and you can do both: answers typed here become the first set of answers, and you can add more in the app afterwards."),
    ("b", "Question ID — leave it blank for a new question. On a questionnaire you downloaded back out of the app it is already filled in; DO NOT EDIT OR DELETE IT. It is how the app knows that row is the same question you have already collected answers against."),
    ("b", "You can add columns of your own, move columns around, and put a title above the headings. The app finds the columns by their headings, not by where they are."),
    ("h", "Editing a questionnaire people have already answered"),
    ("p", "Once an answer has been recorded against a question, that question stops being editable — an answer only means anything next to the words it was given under."),
    ("b", "Change the wording of an answered question and the app keeps the old question and its answers, and adds your new wording as a new question underneath. Nothing is lost and no recorded answer changes meaning."),
    ("b", "Delete a row for an answered question and the app retires it instead: it stops being asked, and its answers stay in the record."),
    ("b", "Questions nobody has answered yet can be reworded, reordered and deleted freely."),
    ("h", "Sending your questions to another designer"),
    ("p", "Two different files come out of a questionnaire, and only one of them is meant to be passed on."),
    ("b", "QUESTION SET — the questions, their order, their help text and their Required flags, and nothing else: no answers, no respondents' names, no sittings. This is the one to send. Whoever receives it uploads it and gets their own empty questionnaire with your questions in it, to run their own fieldwork against."),
    ("b", "FULL WORKBOOK — the same questions PLUS every sitting recorded against them: each respondent's name, their notes and every answer they gave. That is your fieldwork rather than your instrument, and only you, a designer working on the same design workshop, and an admin can download it."),
    ("b", "A workbook that came out of the app and still has answers in it imports its QUESTIONS ONLY. Those answers are already recorded in the platform under the names of the people who recorded them, and copying them into a second questionnaire under your name would duplicate the fieldwork and misattribute it. Answers you typed into a blank pro-forma yourself are imported as normal — that file has no Question IDs in it, which is how the app tells the two apart."),
    ("h", "An example"),
]
_HELP_EXAMPLE = [
    ["Section Code", "Section Title", "Question ID", "Question", "Required", "Answer"],
    ["", "About the craft", "", "How long have you practised this craft?", "Yes", "22 years"],
    ["", "", "", "Who taught you?", "No", ""],
    ["", "Materials", "", "Where do you buy your yarn?", "Yes", ""],
]


def _help_sheet(ws: Worksheet) -> None:
    ws.sheet_properties.tabColor = "FF9CA3AF"
    row = 1
    for kind, text in _HELP_LINES:
        cell = _put(ws, row, 1, text if kind != "b" else f"•  {text}")
        if kind == "h":
            cell.font = Font(bold=True, size=13, color=f"FF{_BRAND}")
            row += 1
        elif kind == "p":
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=6)
            row += 2
        else:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=6)
            row += 3
    row += 1
    for offset, line in enumerate(_HELP_EXAMPLE):
        for column, value in enumerate(line, start=1):
            cell = _put(ws, row + offset, column, value)
            if offset == 0:
                cell.font = Font(bold=True, color=f"FF{_BRAND}")
    ws.column_dimensions["A"].width = 26
    for letter in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[letter].width = 24


def _question_columns(entry_labels: list[str]) -> list[tuple[str, str]]:
    """The working sheet's headings: the fixed ones, then one Answer/Notes pair per sitting.

    The label is repeated in both headings of a pair, which is exactly what the parser reads to pair
    them back up — so a workbook exported with three sittings re-imports as three sittings.
    """
    columns = [
        ("sectionCode", "Section Code"),
        ("sectionTitle", "Section Title"),
        ("questionId", "Question ID"),
        ("prompt", "Question"),
        ("help", "Help text"),
        ("required", "Required"),
    ]
    for label in entry_labels or [""]:
        suffix = f" — {label}" if label else ""
        columns.append(("answer", f"Answer{suffix}"))
        columns.append(("notes", f"Notes{suffix}"))
    return columns


def _write_questions_sheet(
    ws: Worksheet,
    sections: list[dict[str, Any]],
    entry_labels: list[str],
) -> None:
    ws.sheet_properties.tabColor = f"FF{_BRAND}"
    columns = _question_columns(entry_labels)
    _header(ws, 1, columns)

    grey = Font(color="FF6B7280")
    wrap = Alignment(wrap_text=True, vertical="top")
    row = 2
    for section in sections:
        first = True
        for question in section.get("questions") or []:
            values: list[Any] = [
                section.get("code") if first else "",
                section.get("title") if first else "",
                question.get("id") or "",
                question.get("prompt") or "",
                question.get("helpText") or "",
                "Yes" if question.get("isRequired") else "No",
            ]
            answers = question.get("answers") or {}
            notes = question.get("answerNotes") or {}
            for label in entry_labels or [""]:
                values.append(answers.get(label, ""))
                values.append(notes.get(label, ""))
            for index, value in enumerate(values, start=1):
                cell = _put(ws, row, index, value)
                cell.alignment = wrap
                if index == 3 and value:
                    # The identity column. Greyed because it is the one column a designer must not
                    # retype: it is what ties this row to the answers already recorded against it.
                    cell.font = grey
            first = False
            row += 1
        if not (section.get("questions") or []):
            # A section with no questions still has to survive the round trip, or a designer who
            # typed their headings first and their questions second loses the headings.
            _put(ws, row, 1, section.get("code") or "")
            _put(ws, row, 2, section.get("title") or "")
            row += 1


def build_pro_forma() -> bytes:
    """The BLANK pro-forma: headings, a Details sheet to name the questionnaire, and instructions.

    Deliberately EMPTY under the headings. An earlier shape seeded three example questions on the
    working sheet to show what one looked like, which reads well and imports badly: the designer who
    adds their questions below the examples uploads a questionnaire whose first three questions are
    "How long have you practised this craft?" and has to work out where they came from. The worked
    example lives on the instructions sheet, where it cannot be imported.
    """
    wb = Workbook()
    questions = wb.active
    questions.title = SHEET_QUESTIONS
    _write_questions_sheet(questions, [], [])
    _details_sheet(
        wb.create_sheet(title=SHEET_DETAILS),
        title="",
        description="",
        questionnaire_id="",
        version=None,
    )
    _help_sheet(wb.create_sheet(title=SHEET_HELP))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_questionnaire_workbook(
    *,
    title: str,
    description: str | None,
    questionnaire_id: str,
    version: int,
    sections: list[dict[str, Any]],
    entry_labels: list[str] | None = None,
) -> bytes:
    """An EXISTING questionnaire as the same workbook — the download half of the edit loop.

    ``sections`` is ``[{code, title, questions: [{id, prompt, helpText, isRequired, answers,
    answerNotes}]}]``, where ``answers``/``answerNotes`` are ``{sitting label: text}``. Question ids
    are written into the Question ID column, which is what makes a download-edit-upload round trip
    an EDIT of these questions rather than a second copy of them.
    """
    labels = list(entry_labels or [])
    wb = Workbook()
    questions = wb.active
    questions.title = SHEET_QUESTIONS
    _write_questions_sheet(questions, sections, labels)
    _details_sheet(
        wb.create_sheet(title=SHEET_DETAILS),
        title=_sanitise(title or ""),
        description=_sanitise(description or ""),
        questionnaire_id=questionnaire_id,
        version=version,
    )
    _help_sheet(wb.create_sheet(title=SHEET_HELP))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


QUESTION_SET_CONTENTS = "Questions only — no answers, no respondents' names, no recorded sittings."

_QUESTION_SET_NOTE = (
    "This is a QUESTION SET: one questionnaire's questions and nothing else. It carries no answers, "
    "no respondents' names and no recorded sittings, which is what makes it safe to send to another "
    "designer. 'Questionnaire ID' is blank and the Question ID column is empty ON PURPOSE — "
    "uploading this file creates a NEW questionnaire that belongs to you, rather than editing the "
    "one it came from. Type your own answers into the Answer column, or leave it empty and record "
    "them in the app."
)


def build_question_set_workbook(
    *,
    title: str,
    description: str | None,
    sections: list[dict[str, Any]],
    source_title: str | None = None,
    shared_by: str | None = None,
    exported_on: str | None = None,
) -> bytes:
    """One questionnaire's QUESTIONS, as a workbook a designer may hand to another designer.

    ================================================================================================
    WHY THIS EXISTS AS A SECOND ARTEFACT RATHER THAN AS A RELAXED PERMISSION ON THE FIRST
    ================================================================================================

    ``build_questionnaire_workbook`` is deliberately LOSSLESS — every sitting, every respondent's
    name, every answer, every retired question — because a download-edit-upload round trip has to be.
    That is why its route refuses anyone but the owner, a designer on its design workshop, or an
    admin, and that refusal is correct: the artefact really does carry somebody's fieldwork.

    But the thing designers actually want to send each other is the INSTRUMENT — "here are the
    eighteen questions we ask weavers, use them". There was no way to produce that, so sharing was
    impossible and the only workaround was to widen the gate on a file full of respondents' names.
    This is the second artefact instead. It is not the same file with columns blanked out: it is
    built from a query that never reads an answer at all (see ``load_question_set`` in
    ``services/questionnaire_forms.py``), so it cannot leak one by omission.

    THREE THINGS ARE DELIBERATELY ABSENT AND EACH OF THEM WOULD CAUSE A DISTINCT BUG IF ADDED BACK:

    * **No Answer/Notes values.** The obvious one, and the point of the artefact.
    * **No Question IDs.** Those ids belong to the SENDER's questionnaire. Left in, a receiving
      designer's re-upload would report every row as "Question ID … does not belong to this
      questionnaire", and — worse — the ids are the signal the import uses to recognise a workbook
      that came out of the app and therefore to refuse to re-record its answers (see
      ``create_from_parsed``). A shared question set has to read as what it is: a filled-in
      pro-forma.
    * **No Questionnaire ID on the Details sheet.** Same reason from the other side: with one, the
      receiving designer's ``POST /questionnaires/{id}/upload`` answers 409 "that workbook was
      downloaded from a different questionnaire", which is exactly right for the full workbook and
      exactly wrong for a question set they are entitled to use.

    RETIRED QUESTIONS ARE NOT INCLUDED, and that is the one place this differs from the full
    workbook for a reason other than privacy. A retired question is not part of the instrument any
    more — it is kept only because answers hang off it. Sending it would plant a question the sender
    deliberately replaced into the receiver's brand-new form, next to its replacement.
    """
    extra = [("Contents", QUESTION_SET_CONTENTS)]
    if source_title:
        extra.append(("Exported from", source_title))
    if shared_by:
        extra.append(("Shared by", shared_by))
    if exported_on:
        extra.append(("Exported on", exported_on))

    wb = Workbook()
    questions = wb.active
    questions.title = SHEET_QUESTIONS
    # `entry_labels=[]` gives ONE blank Answer/Notes pair, exactly as the pro-forma does, so the
    # receiving designer can record on paper in the same file. It carries no values because every
    # question handed in below has empty `answers`/`answerNotes`.
    _write_questions_sheet(questions, sections, [])
    _details_sheet(
        wb.create_sheet(title=SHEET_DETAILS),
        title=_sanitise(title or ""),
        description=_sanitise(description or ""),
        questionnaire_id="",
        version=None,
        extra=[(label, _sanitise(value)) for label, value in extra],
        note=_QUESTION_SET_NOTE,
    )
    _help_sheet(wb.create_sheet(title=SHEET_HELP))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _filename_stem(title: str | None) -> str:
    stem = re.sub(r"[^A-Za-z0-9 _-]+", "", _sanitise(title or "questionnaire")).strip()
    return re.sub(r"\s+", "-", stem)[:80].strip("-") or "questionnaire"


def download_filename(title: str | None) -> str:
    """A safe, readable filename for a questionnaire download."""
    return f"{_filename_stem(title)}.xlsx"


def question_set_filename(title: str | None) -> str:
    """The same, for the questions-only download.

    The ``-questions`` suffix is not decoration. Both downloads land in the same Downloads folder
    with the same questionnaire title on them, and the two files are the difference between sending
    a colleague your question list and sending them every respondent you have ever interviewed. The
    name is the last thing standing between a designer and that mistake, so it says which one this
    is.
    """
    return f"{_filename_stem(title)}-questions.xlsx"
