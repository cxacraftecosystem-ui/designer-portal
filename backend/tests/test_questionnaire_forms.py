"""Custom questionnaires: the .xlsx round trip, and what an edit does to answers already recorded.

Two things are pinned here, and both of them are ways a designer silently loses work.

**THE ROUND TRIP MUST BE LOSSLESS.** Generate a pro-forma, fill it in, parse it, write it back out,
parse it again — the same sections, the same questions, in the same order, with the same answers. It
is the only property that makes edit-in-Excel safe to offer: a designer who downloads their
questionnaire, changes one word and re-uploads is trusting that the other thirty-nine questions come
back exactly as they went in. If a question can be lost in either direction, this feature quietly
deletes fieldwork every time somebody uses it as intended.

**AN ANSWER MUST NEVER CHANGE MEANING.** A question answered "12" under "How many looms do you own?"
must not end up sitting under "How many weavers work with you?" because the designer reworded it a
week later. Nobody edits an answer, nothing in any log says anything happened, and the number goes
into a report submitted to a ministry. The rule (see services/questionnaire_forms.py) is that
rewording an ANSWERED question supersedes it rather than overwriting it, and deleting one retires it
rather than orphaning its answers — and the last test in this file is the one that would catch a
future refactor that "simplified" that away.

The parser tests need nothing but openpyxl. The rule tests need Postgres, because the rule is a
statement about rows and the ``ON DELETE RESTRICT`` under it is a statement about the database, so
that half skips itself when ``DATABASE_URL`` is not local — exactly as test_designer_roster does.

    docker compose up -d postgres
    cd backend && .venv/Scripts/python.exe -m prisma migrate deploy --schema prisma/schema.prisma
"""

import os
import uuid
import zipfile
from io import BytesIO
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from app.core.db import db
from app.core.security import create_access_token
from app.services.questionnaire_xlsx import (
    QuestionnaireXlsxError,
    build_pro_forma,
    build_questionnaire_workbook,
    derive_section_code,
    parse_questionnaire_workbook,
)

_URL = os.environ.get("DATABASE_URL", "")
_LOCAL = any(host in _URL for host in ("localhost", "127.0.0.1"))

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# The form used everywhere below. Two sections, one of them with an answer already recorded against
# it, because "a questionnaire that has been answered" is the state every interesting rule is about.
FORM = [
    {
        "code": "CRAFT",
        "title": "About the craft",
        "questions": [
            {
                "id": "q-looms",
                "prompt": "How many looms do you own?",
                "helpText": "Count working looms only",
                "isRequired": True,
                "answers": {"Ramesh": "12"},
                "answerNotes": {"Ramesh": "answered through a translator"},
            },
            {"id": "q-taught", "prompt": "Who taught you?", "isRequired": False, "answers": {}},
        ],
    },
    {
        "code": "MAT",
        "title": "Materials",
        "questions": [
            {
                "id": "q-yarn",
                "prompt": "Where do you buy your yarn?",
                "isRequired": True,
                "answers": {"Ramesh": "Panipat mandi"},
            }
        ],
    },
]


def _flatten(parsed: Any) -> list[tuple[str, str, str]]:
    """(section title, prompt, first answer) for every question, in order."""
    return [
        (section.title, question.prompt, next(iter(question.answers.values()), ""))
        for section in parsed.sections
        for question in section.questions
    ]


def _sheet(rows: list[list[Any]], *, title: str = "Questionnaire") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for row in rows:
        ws.append(row)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------------------
# 1. The workbook itself
# --------------------------------------------------------------------------------------


def test_the_pro_forma_is_a_workbook_excel_will_open_without_offering_to_repair_it():
    """An .xlsx is a zip of XML parts and Excel offers to recover the file the moment one of them is
    not well-formed. Eyeballing the download cannot catch that — the workbook still opens, it just
    opens *repaired*. So the archive is unpacked and every part is parsed."""
    import xml.etree.ElementTree as ET

    payload = build_pro_forma()
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        assert archive.testzip() is None
        names = archive.namelist()
        assert "xl/workbook.xml" in names
        for name in names:
            if name.endswith((".xml", ".rels")):
                ET.fromstring(archive.read(name))  # raises if the part is not well-formed


def test_the_pro_forma_carries_no_example_questions_to_import_by_accident():
    """The worked example lives on the instructions sheet, which is never parsed.

    An earlier shape seeded three examples on the WORKING sheet, which reads well and imports badly:
    the designer who types their questions underneath uploads a questionnaire whose first three
    questions are somebody else's, with nothing on screen to say where they came from.
    """
    with pytest.raises(QuestionnaireXlsxError) as excinfo:
        parse_questionnaire_workbook(build_pro_forma(), filename="questionnaire-pro-forma.xlsx")
    assert "no questions under it" in str(excinfo.value)

    book = load_workbook(BytesIO(build_pro_forma()))
    assert book["Questionnaire"].max_row == 1  # the header, and nothing else


def test_a_file_that_is_not_a_workbook_is_refused_with_a_sentence_a_designer_can_act_on():
    """Not a stack trace, and not "invalid file": the message has to name the remedy, because the
    overwhelmingly likely cause is a .xls or a Numbers export and the fix is one Save As away."""
    with pytest.raises(QuestionnaireXlsxError) as excinfo:
        parse_questionnaire_workbook(b"this is a .csv, not a workbook", filename="answers.xls")
    assert "Save As" in str(excinfo.value)


# --------------------------------------------------------------------------------------
# 2. The round trip — generate -> parse -> the same questions
# --------------------------------------------------------------------------------------


def test_generate_then_parse_returns_exactly_the_same_questions():
    payload = build_questionnaire_workbook(
        title="Panipat loom study",
        description="Baseline for the cluster",
        questionnaire_id="qn-1",
        version=1,
        sections=FORM,
        entry_labels=["Ramesh"],
    )
    parsed = parse_questionnaire_workbook(payload, filename="panipat.xlsx")

    assert parsed.title == "Panipat loom study"
    assert parsed.description == "Baseline for the cluster"
    assert parsed.questionnaireId == "qn-1"
    assert [s.code for s in parsed.sections] == ["CRAFT", "MAT"]
    assert _flatten(parsed) == [
        ("About the craft", "How many looms do you own?", "12"),
        ("About the craft", "Who taught you?", ""),
        ("Materials", "Where do you buy your yarn?", "Panipat mandi"),
    ]
    # The identity column is what makes a re-upload an EDIT rather than a second copy.
    assert [q.questionId for s in parsed.sections for q in s.questions] == [
        "q-looms",
        "q-taught",
        "q-yarn",
    ]
    first = parsed.sections[0].questions[0]
    assert first.isRequired is True
    assert first.helpText == "Count working looms only"
    assert first.answerNotes == {"Ramesh": "answered through a translator"}


def test_the_round_trip_survives_being_run_twice():
    """Parse -> write -> parse again. A round trip that is only lossless once is not lossless: the
    designer who downloads, edits and re-uploads on Monday does it again on Thursday."""
    once = parse_questionnaire_workbook(
        build_questionnaire_workbook(
            title="Loom study",
            description=None,
            questionnaire_id="qn-1",
            version=1,
            sections=FORM,
            entry_labels=["Ramesh"],
        ),
        filename="a.xlsx",
    )
    rewritten = build_questionnaire_workbook(
        title=once.title or "",
        description=once.description,
        questionnaire_id=once.questionnaireId or "",
        version=2,
        sections=[
            {
                "code": section.code,
                "title": section.title,
                "questions": [
                    {
                        "id": q.questionId,
                        "prompt": q.prompt,
                        "helpText": q.helpText,
                        "isRequired": q.isRequired,
                        "answers": q.answers,
                        "answerNotes": q.answerNotes,
                    }
                    for q in section.questions
                ],
            }
            for section in once.sections
        ],
        entry_labels=once.entryLabels,
    )
    twice = parse_questionnaire_workbook(rewritten, filename="a.xlsx")
    assert _flatten(twice) == _flatten(once)
    assert [q.questionId for s in twice.sections for q in s.questions] == [
        q.questionId for s in once.sections for q in s.questions
    ]


def test_several_sittings_round_trip_as_several_sittings():
    """Two answer columns are two people, not one merged column. The heading after the dash is what
    pairs "Answer — Sita" back to "Notes — Sita", so losing it merges two interviews into one."""
    sections = [
        {
            "code": "A",
            "title": "Section A",
            "questions": [
                {
                    "id": "q1",
                    "prompt": "How many looms?",
                    "answers": {"Ramesh": "12", "Sita": "4"},
                    "answerNotes": {"Sita": "shared with her sister"},
                }
            ],
        }
    ]
    parsed = parse_questionnaire_workbook(
        build_questionnaire_workbook(
            title="T",
            description=None,
            questionnaire_id="q",
            version=1,
            sections=sections,
            entry_labels=["Ramesh", "Sita"],
        ),
        filename="t.xlsx",
    )
    question = parsed.sections[0].questions[0]
    assert question.answers == {"Ramesh": "12", "Sita": "4"}
    assert question.answerNotes == {"Sita": "shared with her sister"}
    assert parsed.entryLabels == ["Ramesh", "Sita"]


# --------------------------------------------------------------------------------------
# 3. Forgiveness — and reporting rather than dropping
# --------------------------------------------------------------------------------------


def test_columns_may_be_renamed_reordered_and_pushed_down_the_sheet():
    """The three things a real designer's file does that a positional parser cannot survive: a title
    across the top, columns in their own order, and headings spelled their own way."""
    payload = _sheet(
        [
            ["Kullu shawl questionnaire — draft 3"],
            [],
            ["  ANSWER : ", "section_name", "Questions", "Mandatory?"],
            ["", "About the craft", "", ""],
            ["12", "", "How many looms do you own?", "yes"],
            ["", "", "Who taught you?", "N"],
        ],
        title="Final v3",
    )
    parsed = parse_questionnaire_workbook(payload, filename="kullu.xlsx")
    assert parsed.sheet == "Final v3"
    assert _flatten(parsed) == [
        ("About the craft", "How many looms do you own?", "12"),
        ("About the craft", "Who taught you?", ""),
    ]
    assert parsed.sections[0].questions[0].isRequired is True
    assert parsed.sections[0].questions[1].isRequired is False


def test_a_section_named_once_carries_down_the_rows_beneath_it():
    """How a person lays a table out. Repeating the section on every row must work too, and must not
    produce four sections all called "Materials"."""
    parsed = parse_questionnaire_workbook(
        _sheet(
            [
                ["Section Title", "Question"],
                ["Materials", "Where do you buy yarn?"],
                ["", "What dye do you use?"],
                ["Materials", "Who supplies your mordant?"],
                ["Tools", "How old is your loom?"],
            ]
        ),
        filename="x.xlsx",
    )
    assert [(s.title, len(s.questions)) for s in parsed.sections] == [("Materials", 3), ("Tools", 1)]


def test_a_question_before_any_section_is_kept_and_reported_never_dropped():
    """The commonest malformed file there is: somebody started typing questions and never added a
    heading. Filing them under a default section is right; losing them is not."""
    parsed = parse_questionnaire_workbook(
        _sheet([["Question"], ["How many looms do you own?"], ["Who taught you?"]]),
        filename="x.xlsx",
    )
    assert len(parsed.sections) == 1
    assert len(parsed.sections[0].questions) == 2
    warnings = [p for p in parsed.problems if p.severity == "warning"]
    assert warnings and warnings[0].row == 2
    assert "General" in warnings[0].reason


def test_an_unreadable_required_value_is_reported_with_its_excel_row_number():
    """"maybe" is not yes or no. The question is still imported — losing it over one cell would be
    absurd — and the designer is told which row to look at, by the number Excel's gutter shows."""
    parsed = parse_questionnaire_workbook(
        _sheet(
            [
                ["Section Title", "Question", "Required"],
                ["Craft", "How many looms?", "maybe"],
            ]
        ),
        filename="x.xlsx",
    )
    assert len(parsed.sections[0].questions) == 1
    problem = parsed.problems[0]
    assert problem.row == 2
    assert "maybe" in problem.reason
    assert problem.severity == "warning"


def test_a_formula_with_no_cached_result_is_reported_as_a_formula_not_as_a_blank_row():
    """THE ONE CASE openpyxl cannot help with. A workbook written by anything other than Excel has
    formulas but no cached results, so the cell reads as empty. Telling a designer "row 3 is blank"
    about a row that visibly says ``=B1&" 2024"`` on their screen is the most confusing thing this
    parser could say."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questionnaire"
    ws.append(["Section Title", "Question"])
    ws.append(["Craft", "How many looms?"])
    ws["A3"] = "=A2"
    ws["B3"] = '=B2&" (2024)"'
    buffer = BytesIO()
    wb.save(buffer)

    parsed = parse_questionnaire_workbook(buffer.getvalue(), filename="x.xlsx")
    assert len(parsed.sections[0].questions) == 1
    errors = [p for p in parsed.problems if p.severity == "error"]
    assert errors, [p.payload() for p in parsed.problems]
    assert errors[0].row == 3
    assert "formula" in errors[0].reason


def test_a_workbook_with_no_question_column_is_refused_rather_than_imported_empty():
    with pytest.raises(QuestionnaireXlsxError) as excinfo:
        parse_questionnaire_workbook(
            _sheet([["Name", "Village", "Phone"], ["Ramesh", "Panipat", "0"]]), filename="x.xlsx"
        )
    assert "'Question' column" in str(excinfo.value)


def test_a_section_code_is_derived_from_the_title_so_inserting_a_section_does_not_renumber():
    """Positional codes look tidier and are wrong: inserting a section at the top would renumber
    every section below it, so a re-upload matches nothing and the whole form churns."""
    taken: set[str] = set()
    assert derive_section_code("About the craft", taken) == "ABOUT_THE_CRAFT"
    assert derive_section_code("About the craft", taken) == "ABOUT_THE_CRAFT_2"
    assert derive_section_code("", taken) == "SECTION"


# --------------------------------------------------------------------------------------
# 4. The edit-after-answers rule. Needs Postgres.
# --------------------------------------------------------------------------------------

pytestmark_db = [
    pytest.mark.skipif(
        not _LOCAL, reason="needs a LOCAL database; refuses to run against a remote DATABASE_URL"
    ),
    pytest.mark.anyio,
]


@pytest.fixture(scope="module")
def anyio_backend():
    return "asyncio"


@pytest.fixture(scope="module")
async def world():
    """A designer, an admin, and a live TestClient.

    Rows are created here rather than inside a test because the Prisma client is shared with the
    running app and bound to the TestClient's event loop; touching it from a test's own loop is the
    kind of cross-loop use that fails intermittently rather than honestly. Addresses carry a per-run
    stamp because ``User.email`` is unique and a fixed address passes once and fails for ever after.
    """
    if not _LOCAL:
        pytest.skip("needs a local database")
    from fastapi.testclient import TestClient

    from app.main import app

    stamp = uuid.uuid4().hex[:8]
    people: dict[str, Any] = {}
    await db.connect()
    try:
        for slug, role, name in (
            ("designer", "DESIGNER", "Custom Questionnaire Designer"),
            ("colleague", "DESIGNER", "Another Designer"),
            # A THIRD designer, who owns nothing and recorded nothing. The authorship tests need
            # somebody who is neither the form's owner nor the sitting's author; with only two
            # designers every refusal could be explained by ownership of the form alone.
            ("stranger", "DESIGNER", "Unrelated Designer"),
            ("admin", "ADMIN", "Questionnaire Admin"),
        ):
            people[slug] = await db.user.create(
                data={
                    "email": f"cq-{slug}-{stamp}@example.org",
                    "name": name,
                    "role": role,
                }
            )
    finally:
        await db.disconnect()
    with TestClient(app) as client:
        yield {"client": client, "people": people, "stamp": stamp}


@pytest.fixture
def client(world):
    return world["client"]


def _headers(world: dict[str, Any], slug: str = "designer") -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token(subject=world['people'][slug].id)}"}


def _upload(client: Any, world: dict[str, Any], payload: bytes, name: str = "form.xlsx") -> Any:
    return client.post(
        "/api/questionnaires/upload",
        files={"file": (name, payload, XLSX_MIME)},
        headers=_headers(world),
    )


def _questions(form: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """Every question in the form, keyed by prompt."""
    return {q["prompt"]: q for section in form["sections"] for q in section["questions"]}


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_uploading_a_filled_in_pro_forma_creates_a_questionnaire_with_its_answers(
    client, world
):
    """A spreadsheet that ARRIVES with answers in it must produce the same rows as answers typed
    into the app, or every downstream reader needs a special case for each kind."""
    payload = build_questionnaire_workbook(
        title=f"Loom study {world['stamp']}",
        description=None,
        questionnaire_id="",
        version=1,
        sections=[
            {
                "code": "CRAFT",
                "title": "About the craft",
                "questions": [
                    {"id": "", "prompt": "How many looms do you own?", "answers": {"Ramesh": "12"}},
                    {"id": "", "prompt": "Who taught you?", "answers": {}},
                ],
            }
        ],
        entry_labels=["Ramesh"],
    )
    response = _upload(client, world, payload)
    assert response.status_code == 201, response.text
    body = response.json()
    form = body["questionnaire"]

    assert form["questionCount"] == 2
    assert body["report"]["created"] == 2
    assert body["report"]["problems"] == []
    # The uploaded answers are an ordinary sitting, marked as having come off the spreadsheet.
    assert len(form["entries"]) == 1
    assert form["entries"][0]["source"] == "UPLOAD"
    assert form["entries"][0]["respondentName"] == "Ramesh"
    answers = {a["questionId"]: a["answerText"] for a in form["entries"][0]["answers"]}
    looms = _questions(form)["How many looms do you own?"]
    assert answers[looms["id"]] == "12"
    assert looms["hasAnswers"] is True
    assert _questions(form)["Who taught you?"]["hasAnswers"] is False


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_rewording_an_answered_question_supersedes_it_and_the_answer_keeps_its_words(
    client, world
):
    """THE TEST THIS FILE EXISTS FOR.

    "How many looms do you own?" is answered "12". The designer reworks the form and the question
    becomes "How many weavers work with you?". If the prompt were simply overwritten, the repository
    would now assert that this artisan works with twelve weavers — nobody edited an answer, nothing
    says anything happened, and the number goes into a ministry report.

    What must happen instead: the original question is retired WITH ITS ORIGINAL WORDING and keeps
    its answer, the new wording is stored as a NEW question, and the two are linked.
    """
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Supersede {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[
                {
                    "code": "CRAFT",
                    "title": "About the craft",
                    "questions": [
                        {"id": "", "prompt": "How many looms do you own?", "answers": {"R": "12"}}
                    ],
                }
            ],
            entry_labels=["R"],
        ),
    ).json()
    questionnaire_id = created["questionnaire"]["id"]
    original = _questions(created["questionnaire"])["How many looms do you own?"]

    response = client.patch(
        f"/api/questionnaires/{questionnaire_id}/questions/{original['id']}",
        json={"prompt": "How many weavers work with you?"},
        headers=_headers(world),
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["action"] == "superseded"
    assert body["questionId"] == original["id"]
    replacement_id = body["replacementId"]
    assert replacement_id != original["id"]

    form = body["questionnaire"]
    by_id = {q["id"]: q for section in form["sections"] for q in section["questions"]}
    # The answered question keeps the words it was answered under, and stops being asked.
    assert by_id[original["id"]]["prompt"] == "How many looms do you own?"
    assert by_id[original["id"]]["isActive"] is False
    assert by_id[original["id"]]["retiredAt"] is not None
    assert by_id[original["id"]]["supersededById"] == replacement_id
    # The correction the designer asked for exists, and has no answers of its own.
    assert by_id[replacement_id]["prompt"] == "How many weavers work with you?"
    assert by_id[replacement_id]["isActive"] is True
    assert by_id[replacement_id]["hasAnswers"] is False
    # And the "12" is still attached to the question about looms.
    answers = {a["questionId"]: a["answerText"] for a in form["entries"][0]["answers"]}
    assert answers[original["id"]] == "12"
    assert replacement_id not in answers
    assert form["version"] == created["questionnaire"]["version"] + 1


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_an_unanswered_question_is_reworded_in_place_with_no_ceremony(client, world):
    """The ordinary case — a designer drafting their form — must stay frictionless. Superseding an
    unanswered question would litter every draft with retired rows nobody ever answered."""
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Draft {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[{"code": "A", "title": "A", "questions": [{"id": "", "prompt": "Typo?"}]}],
        ),
    ).json()
    questionnaire_id = created["questionnaire"]["id"]
    question = _questions(created["questionnaire"])["Typo?"]

    body = client.patch(
        f"/api/questionnaires/{questionnaire_id}/questions/{question['id']}",
        json={"prompt": "Fixed."},
        headers=_headers(world),
    ).json()
    assert body["action"] == "updated"
    form = body["questionnaire"]
    assert [q["prompt"] for s in form["sections"] for q in s["questions"]] == ["Fixed."]
    assert form["version"] == created["questionnaire"]["version"]


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_deleting_an_answered_question_retires_it_and_keeps_the_answer(client, world):
    """Deleting a question orphans its answers, so an answered one is never deleted. An unanswered
    one is, because there is nothing to orphan and a form littered with every question its author
    thought better of is not a form."""
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Retire {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[
                {
                    "code": "A",
                    "title": "A",
                    "questions": [
                        {"id": "", "prompt": "Answered question", "answers": {"R": "yes"}},
                        {"id": "", "prompt": "Unanswered question"},
                    ],
                }
            ],
            entry_labels=["R"],
        ),
    ).json()
    questionnaire_id = created["questionnaire"]["id"]
    answered = _questions(created["questionnaire"])["Answered question"]
    unanswered = _questions(created["questionnaire"])["Unanswered question"]

    body = client.delete(
        f"/api/questionnaires/{questionnaire_id}/questions/{answered['id']}",
        headers=_headers(world),
    ).json()
    assert body["action"] == "retired"
    retired = {q["id"]: q for s in body["questionnaire"]["sections"] for q in s["questions"]}
    assert retired[answered["id"]]["isActive"] is False
    answers = {a["questionId"] for a in body["questionnaire"]["entries"][0]["answers"]}
    assert answered["id"] in answers  # the answer survived the delete

    body = client.delete(
        f"/api/questionnaires/{questionnaire_id}/questions/{unanswered['id']}",
        headers=_headers(world),
    ).json()
    assert body["action"] == "deleted"
    remaining = {q["id"] for s in body["questionnaire"]["sections"] for q in s["questions"]}
    assert unanswered["id"] not in remaining


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_retired_question_cannot_collect_new_answers(client, world):
    """A wording the designer deliberately replaced must not quietly carry on gathering evidence."""
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"NoNewAnswers {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[
                {
                    "code": "A",
                    "title": "A",
                    "questions": [{"id": "", "prompt": "Original wording", "answers": {"R": "12"}}],
                }
            ],
            entry_labels=["R"],
        ),
    ).json()
    questionnaire_id = created["questionnaire"]["id"]
    original = _questions(created["questionnaire"])["Original wording"]
    client.patch(
        f"/api/questionnaires/{questionnaire_id}/questions/{original['id']}",
        json={"prompt": "New wording"},
        headers=_headers(world),
    )
    entry = client.post(
        f"/api/questionnaires/{questionnaire_id}/entries",
        json={"respondentName": "Sita"},
        headers=_headers(world),
    ).json()
    response = client.put(
        f"/api/questionnaires/{questionnaire_id}/entries/{entry['id']}/answers",
        json={"answers": [{"questionId": original["id"], "answerText": "4"}]},
        headers=_headers(world),
    )
    assert response.status_code == 422, response.text
    assert "retired" in response.json()["detail"]


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_re_uploading_an_edited_workbook_supersedes_answered_and_retires_removed(
    client, world
):
    """The same rule, applied by the spreadsheet path rather than the editor.

    The two paths must not drift into two different answers to the same question, so this asserts
    the whole shape at once: an untouched question stays put, a reworded ANSWERED one is superseded,
    a reworded UNANSWERED one is edited in place, one deleted from the sheet is retired because it
    has answers, and a brand new row is created.
    """
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Reupload {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[
                {
                    "code": "A",
                    "title": "Section A",
                    "questions": [
                        {"id": "", "prompt": "Untouched question", "answers": {"R": "a"}},
                        {"id": "", "prompt": "Answered and reworded", "answers": {"R": "12"}},
                        {"id": "", "prompt": "Unanswered and reworded"},
                        {"id": "", "prompt": "Removed but answered", "answers": {"R": "c"}},
                    ],
                }
            ],
            entry_labels=["R"],
        ),
    ).json()
    questionnaire_id = created["questionnaire"]["id"]
    before = _questions(created["questionnaire"])

    # Download it back — the ids in that file are what make this an edit rather than a second copy.
    download = client.get(
        f"/api/questionnaires/{questionnaire_id}/xlsx", headers=_headers(world)
    )
    assert download.status_code == 200, download.text
    assert download.headers["content-type"].startswith(XLSX_MIME)
    roundtripped = parse_questionnaire_workbook(download.content, filename="edit.xlsx")
    assert roundtripped.questionnaireId == questionnaire_id

    edited = []
    for section in roundtripped.sections:
        questions = []
        for question in section.questions:
            if question.prompt == "Removed but answered":
                continue  # the designer deleted this row
            prompt = question.prompt
            if prompt in ("Answered and reworded", "Unanswered and reworded"):
                prompt = prompt + " (v2)"
            questions.append(
                {
                    "id": question.questionId,
                    "prompt": prompt,
                    "helpText": question.helpText,
                    "isRequired": question.isRequired,
                    "answers": question.answers,
                    "answerNotes": question.answerNotes,
                }
            )
        questions.append({"id": "", "prompt": "Brand new question"})
        edited.append({"code": section.code, "title": section.title, "questions": questions})

    response = client.post(
        f"/api/questionnaires/{questionnaire_id}/upload",
        files={
            "file": (
                "edit.xlsx",
                build_questionnaire_workbook(
                    title=roundtripped.title or "",
                    description=None,
                    questionnaire_id=questionnaire_id,
                    version=2,
                    sections=edited,
                    entry_labels=roundtripped.entryLabels,
                ),
                XLSX_MIME,
            )
        },
        headers=_headers(world),
    )
    assert response.status_code == 200, response.text
    report = response.json()["report"]
    form = response.json()["questionnaire"]
    by_id = {q["id"]: q for s in form["sections"] for q in s["questions"]}

    assert report["superseded"] == 1
    assert report["retired"] == 1
    assert report["created"] == 1

    # Untouched: still there, still active, still holding its answer.
    assert by_id[before["Untouched question"]["id"]]["isActive"] is True
    # Answered + reworded: the original survives with its original words.
    old = by_id[before["Answered and reworded"]["id"]]
    assert old["prompt"] == "Answered and reworded"
    assert old["isActive"] is False
    assert by_id[old["supersededById"]]["prompt"] == "Answered and reworded (v2)"
    # Unanswered + reworded: edited in place, no retired twin left behind.
    assert by_id[before["Unanswered and reworded"]["id"]]["prompt"] == "Unanswered and reworded (v2)"
    assert by_id[before["Unanswered and reworded"]["id"]]["isActive"] is True
    # Removed from the sheet but answered: retired, not deleted, and its answer is still there.
    removed = by_id[before["Removed but answered"]["id"]]
    assert removed["isActive"] is False
    kept_answers = {a["questionId"] for a in form["entries"][0]["answers"]}
    assert before["Removed but answered"]["id"] in kept_answers
    # And the form gained the new question.
    assert "Brand new question" in {q["prompt"] for q in by_id.values()}


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_downloading_and_re_uploading_unchanged_does_not_resurrect_retired_questions(
    client, world
):
    """A REGRESSION, found by reading the export and the importer next to each other.

    The export deliberately writes retired questions out, because dropping them would lose the
    answers recorded against them. So every download carries those rows, ids and all, and every
    re-upload matches them. An importer that treats "present in the file" as "the designer wants
    this back" therefore resurrects every question anybody has ever replaced — each one sitting next
    to the replacement that superseded it — on a round trip that changed NOTHING. Downloading a
    questionnaire and uploading it straight back must be a no-op.
    """
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Noop {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[
                {
                    "code": "A",
                    "title": "A",
                    "questions": [{"id": "", "prompt": "Original wording", "answers": {"R": "12"}}],
                }
            ],
            entry_labels=["R"],
        ),
    ).json()["questionnaire"]
    original = _questions(created)["Original wording"]
    superseded = client.patch(
        f"/api/questionnaires/{created['id']}/questions/{original['id']}",
        json={"prompt": "Replacement wording"},
        headers=_headers(world),
    ).json()
    replacement_id = superseded["replacementId"]
    version_after_supersede = superseded["questionnaire"]["version"]

    # Download and upload back, byte for byte.
    same_file = client.get(
        f"/api/questionnaires/{created['id']}/xlsx", headers=_headers(world)
    ).content
    response = client.post(
        f"/api/questionnaires/{created['id']}/upload",
        files={"file": ("same.xlsx", same_file, XLSX_MIME)},
        headers=_headers(world),
    )
    assert response.status_code == 200, response.text
    report = response.json()["report"]
    form = response.json()["questionnaire"]
    by_id = {q["id"]: q for s in form["sections"] for q in s["questions"]}

    assert report["created"] == 0
    assert report["superseded"] == 0
    assert report["retired"] == 0
    assert report["removed"] == 0
    # The retired question is still retired, still superseded by the same replacement.
    assert by_id[original["id"]]["isActive"] is False
    assert by_id[original["id"]]["supersededById"] == replacement_id
    assert by_id[replacement_id]["isActive"] is True
    # Exactly two questions — no third one conjured out of the round trip.
    assert len(by_id) == 2
    assert form["version"] == version_after_supersede


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_uploading_a_workbook_downloaded_from_a_different_questionnaire_is_refused(
    client, world
):
    """A designer picking the wrong file out of their downloads folder would otherwise retire this
    questionnaire's ENTIRE question set in one press, as "absent from the upload"."""
    first = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"First {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[{"code": "A", "title": "A", "questions": [{"id": "", "prompt": "Q one"}]}],
        ),
    ).json()["questionnaire"]
    second = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Second {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[{"code": "B", "title": "B", "questions": [{"id": "", "prompt": "Q two"}]}],
        ),
    ).json()["questionnaire"]

    wrong_file = client.get(
        f"/api/questionnaires/{second['id']}/xlsx", headers=_headers(world)
    ).content
    response = client.post(
        f"/api/questionnaires/{first['id']}/upload",
        files={"file": ("wrong.xlsx", wrong_file, XLSX_MIME)},
        headers=_headers(world),
    )
    assert response.status_code == 409, response.text
    assert second["id"] in response.json()["detail"]
    # And the first questionnaire is untouched.
    still = client.get(f"/api/questionnaires/{first['id']}", headers=_headers(world)).json()
    assert [q["prompt"] for s in still["sections"] for q in s["questions"]] == ["Q one"]


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_answers_are_recorded_in_the_app_and_saving_twice_changes_nothing(client, world):
    """The other half of the requirement: a questionnaire uploaded EMPTY must still be answerable in
    the platform. Idempotence is what makes a client safe to retry on a flaky connection."""
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Answer in app {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[
                {
                    "code": "A",
                    "title": "A",
                    "questions": [{"id": "", "prompt": "How many looms?"}, {"id": "", "prompt": "Who taught you?"}],
                }
            ],
        ),
    ).json()["questionnaire"]
    assert created["entries"] == []

    entry = client.post(
        f"/api/questionnaires/{created['id']}/entries",
        json={"respondentName": "Ramesh"},
        headers=_headers(world),
    ).json()
    questions = _questions(created)
    batch = {
        "answers": [
            {"questionId": questions["How many looms?"]["id"], "answerText": "12"},
            {"questionId": questions["Who taught you?"]["id"], "answerText": "My mother"},
        ]
    }
    first = client.put(
        f"/api/questionnaires/{created['id']}/entries/{entry['id']}/answers",
        json=batch,
        headers=_headers(world),
    )
    assert first.status_code == 200, first.text
    assert first.json()["saved"] == {"created": 2, "updated": 0, "unchanged": 0}

    again = client.put(
        f"/api/questionnaires/{created['id']}/entries/{entry['id']}/answers",
        json=batch,
        headers=_headers(world),
    )
    assert again.json()["saved"] == {"created": 0, "updated": 0, "unchanged": 2}
    assert len(again.json()["answers"]) == 2


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_questionnaire_attaches_to_a_workshop_and_shows_up_in_the_dropdown(client, world):
    """The dropdown the owner asked for. ``/options`` is its own endpoint because a dropdown that
    silently stops at page one is a designer who cannot find what they uploaded this morning."""
    workshop = client.post(
        "/api/design-workshops",
        json={"title": f"Attach test {world['stamp']}"},
        headers=_headers(world),
    )
    assert workshop.status_code == 201, workshop.text
    workshop_id = workshop.json()["id"]

    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Attachable {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[{"code": "A", "title": "A", "questions": [{"id": "", "prompt": "Q?"}]}],
        ),
    ).json()["questionnaire"]

    patched = client.patch(
        f"/api/questionnaires/{created['id']}",
        json={"designWorkshopId": workshop_id},
        headers=_headers(world),
    )
    assert patched.status_code == 200, patched.text
    assert patched.json()["designWorkshopId"] == workshop_id

    scoped = client.get(
        "/api/questionnaires",
        params={"designWorkshopId": workshop_id},
        headers=_headers(world),
    ).json()
    assert [row["id"] for row in scoped["items"]] == [created["id"]]

    options = client.get("/api/questionnaires/options", headers=_headers(world)).json()
    assert created["id"] in {row["id"] for row in options}

    # Detaching is sending null, not omitting the field.
    detached = client.patch(
        f"/api/questionnaires/{created['id']}",
        json={"designWorkshopId": None},
        headers=_headers(world),
    ).json()
    assert detached["designWorkshopId"] is None


def _annexure_blocks(blocks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Everything from the questionnaire annexure's heading to the end of the document."""
    for index, block in enumerate(blocks):
        if block["type"] == "HEADING" and "Questionnaire responses" in _runs_text(block["runs"]):
            return blocks[index:]
    return []


def _runs_text(runs: list[dict[str, Any]]) -> str:
    return "".join(run["text"] for run in runs)


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_the_report_annexure_prints_a_two_section_form_section_by_section(client, world):
    """THE ANNEXURE, THROUGH THE ROUTE A DESIGNER ACTUALLY PRESSES, ON A FORM WITH TWO SECTIONS.

    The annexure's own test file builds its ``QuestionnaireItem``s by hand, so nothing there ever
    reached ``report_items`` — the half that turns five tables into the list the document prints.
    That half had the bug this test exists for. ``QuestionnaireFormQuestion.sortOrder`` is scoped to
    its SECTION (the upload numbers each section's questions from 1), so a flat query ordered by
    sortOrder alone comes back INTERLEAVED — A1, B1, A2, B2 — and the annexure, which starts a fresh
    table every time the section label changes, printed one single-row table per question with the
    label repeated above each, A/B/A/B down the page. Every question of one section separated from
    its neighbours, in an appendix of evidence handed to a ministry officer.

    A ONE-SECTION FIXTURE CANNOT SEE IT, which is exactly why this one has two sections of two
    questions each and asserts the shape of the tables rather than only the presence of the text.

    THROUGH THE PREVIEW ROUTE, not by calling ``report_items``: the Prisma singleton is bound to the
    TestClient's event loop (see the ``world`` fixture), and going in by HTTP walks the whole chain a
    designer's press does — ``load_workshop_or_404`` -> ``_report_inputs`` -> the template's
    ``ANNEXURE_QUESTIONNAIRES`` -> ``attach_report_questionnaires`` -> ``report_items`` -> the
    builder's one branch — rather than the last hop of it.
    """
    workshop = client.post(
        "/api/design-workshops",
        json={"title": f"Annexure order {world['stamp']}"},
        headers=_headers(world),
    )
    assert workshop.status_code == 201, workshop.text
    workshop_id = workshop.json()["id"]

    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Two sections {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[
                {
                    "code": "CRAFT",
                    "title": "About the craft",
                    "questions": [
                        {"id": "", "prompt": "How many looms do you own?",
                         "answers": {"Ramesh": "12"}},
                        {"id": "", "prompt": "Who taught you?",
                         "answers": {"Ramesh": "My father"}},
                    ],
                },
                {
                    "code": "MAT",
                    "title": "Materials",
                    "questions": [
                        {"id": "", "prompt": "Where do you buy your yarn?",
                         "answers": {"Ramesh": "Panipat mandi"}},
                        {"id": "", "prompt": "What does a kilo cost?",
                         "answers": {"Ramesh": "Rs 240"}},
                    ],
                },
            ],
            entry_labels=["Ramesh"],
        ),
    )
    assert created.status_code == 201, created.text
    questionnaire_id = created.json()["questionnaire"]["id"]

    attached = client.patch(
        f"/api/questionnaires/{questionnaire_id}",
        json={"designWorkshopId": workshop_id},
        headers=_headers(world),
    )
    assert attached.status_code == 200, attached.text

    preview = client.get(
        f"/api/design-workshops/{workshop_id}/report/preview", headers=_headers(world)
    )
    assert preview.status_code == 200, preview.text
    annexure = _annexure_blocks(preview.json()["blocks"])
    assert annexure, "the questionnaire annexure never reached the preview at all"

    # The respondent's own heading, and everything under it.
    for index, block in enumerate(annexure):
        if block["type"] == "HEADING" and _runs_text(block["runs"]) == "Ramesh":
            sitting = annexure[index:]
            break
    else:
        raise AssertionError("the sitting's heading is missing from the annexure")

    # The Question|Answer tables, in document order, with the labels that precede them.
    tables = [
        b for b in sitting
        if b["type"] == "TABLE" and [c["header"] for c in b["columns"]] == ["Question", "Answer"]
    ]
    assert len(tables) == 2, (
        "one table per SECTION, not one per question — a table per row is what the interleaved "
        f"ordering produced; got {len(tables)}"
    )
    assert [[_runs_text(row[0]) for row in table["rows"]] for table in tables] == [
        ["How many looms do you own?", "Who taught you?"],
        ["Where do you buy your yarn?", "What does a kilo cost?"],
    ]
    assert [[_runs_text(row[1]) for row in table["rows"]] for table in tables] == [
        ["12", "My father"],
        ["Panipat mandi", "Rs 240"],
    ]

    # Each table is labelled with its own section, once, in the form's order.
    labels = [
        _runs_text(b["runs"]) for b in sitting
        if b["type"] == "PARAGRAPH" and _runs_text(b["runs"]).startswith(("CRAFT", "MAT"))
    ]
    assert labels == ["CRAFT — About the craft", "MAT — Materials"]


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_recorded_answer_is_inside_the_generated_docx_file(client, world):
    """THE ANSWER, IN THE BYTES OF THE FILE A DESIGNER DOWNLOADS. Not a 200, not a block list.

    The test above stops at ``GET /report/preview``, which returns the document as JSON. That is one
    hop short of the deliverable, and the gap is not theoretical: a blank-report defect on this
    project produced a perfectly valid OOXML file — right cover page, openable in Word, nothing in
    it — and only unzipping it revealed that. Every surface upstream of the writer had said fine.

    So this walks the whole chain a designer's press walks — ``POST /design-workshops/{id}/report``
    -> ``_report_inputs`` -> ``ANNEXURE_QUESTIONNAIRES`` in the resolved template ->
    ``attach_report_questionnaires`` -> ``report_items`` -> the builder's one branch ->
    ``report_docx`` — and then OPENS THE ZIP and reads ``word/document.xml``.

    ON PRESENCE, NOT ON LAYOUT. The assertions are that the respondent's name, the question's
    wording and the answer's text are somewhere in the document part. Where they sit, which table
    they are in and how they are styled is the annexure's business and is pinned by the preview test
    above; pinning it again against XML would make every legitimate typographic change fail here.

    The answer strings are nonsense words on purpose: "12" or "Ramesh" could plausibly appear in a
    report for some unrelated reason, and an assertion that can pass by accident proves nothing.
    """
    workshop = client.post(
        "/api/design-workshops",
        json={"title": f"Docx annexure {world['stamp']}"},
        headers=_headers(world),
    )
    assert workshop.status_code == 201, workshop.text
    workshop_id = workshop.json()["id"]

    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Docx instrument {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[{
                "code": "CRAFT",
                "title": "About the craft",
                "questions": [{
                    "id": "",
                    "prompt": "How many looms do you own?",
                    "answers": {"Ramesh Meher": "Twelve looms, ZEBRAFISHLOOM"},
                }],
            }],
            entry_labels=["Ramesh Meher"],
        ),
    )
    assert created.status_code == 201, created.text
    questionnaire_id = created.json()["questionnaire"]["id"]

    attached = client.patch(
        f"/api/questionnaires/{questionnaire_id}",
        json={"designWorkshopId": workshop_id},
        headers=_headers(world),
    )
    assert attached.status_code == 200, attached.text

    generated = client.post(
        f"/api/design-workshops/{workshop_id}/report",
        json={"formats": ["DOCX"], "record": False},
        headers=_headers(world),
    )
    assert generated.status_code == 200, generated.text

    # A .docx is a zip. `word/document.xml` is the body; a file whose cover page is perfect and
    # whose body is empty is exactly the shape of the defect this test exists for.
    with zipfile.ZipFile(BytesIO(generated.content)) as archive:
        document = archive.read("word/document.xml").decode("utf-8")

    for expected in (
        "Questionnaire responses",
        f"Docx instrument {world['stamp']}",
        "Ramesh Meher",
        "How many looms do you own?",
        "ZEBRAFISHLOOM",
    ):
        assert expected in document, (
            f"{expected!r} is not in word/document.xml, so the questionnaire annexure did not "
            "reach the file the designer downloads"
        )


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_an_unanswered_questionnaire_is_never_silently_absent_from_the_report(client, world):
    """AN ATTACHED QUESTIONNAIRE WITH NO ANSWERS MUST SAY SO — ON THE DEFAULT TEMPLATE.

    THE DEFECT, MEASURED. ``x-report-warnings`` was ``"; ".join(warnings)[:900]``. A DCH_STANDARD
    workshop — the default — raises a dozen "required field(s) not recorded" warnings, and the load
    warnings are appended AFTER them, so on this exact fixture the header carried 8 of 12, cut
    mid-word at ``"… 2 required "``, and the questionnaire's own "attached but nothing recorded"
    fell off the end with nothing saying it had. The designer got a report with no questionnaire
    annexure and no sentence anywhere explaining it: the silent omission this whole feature exists
    to end, re-created in the transport that reports on it.

    DCH_STANDARD SPECIFICALLY, and that is the point of the fixture rather than a coin toss. The same
    warning survived on PHOTO_CATALOGUE and COMPACT_SUMMARY throughout, so a test written against a
    short template passes against the broken header.

    WHAT IS PINNED IS THAT THE DESIGNER IS NEVER LEFT WITH NOTHING — the warning itself if it fits,
    and otherwise an explicit count of what did not, pointing at the preview. WHICH warnings win the
    budget when a workshop raises more than fits is an editorial question about every report this
    deployment generates and is deliberately NOT decided here; what is decided is that the loss can
    never again be invisible. The preview half of the assertion is what keeps that honest: the
    warning must exist in full somewhere the designer can actually reach.
    """
    workshop = client.post(
        "/api/design-workshops",
        json={"title": f"Unanswered {world['stamp']}", "templateId": "DCH_STANDARD"},
        headers=_headers(world),
    )
    assert workshop.status_code == 201, workshop.text
    workshop_id = workshop.json()["id"]

    created = client.post(
        "/api/questionnaires",
        json={
            "title": f"Unanswered instrument {world['stamp']}",
            "designWorkshopId": workshop_id,
            "sections": [{
                "title": "About the craft",
                "code": "CRAFT",
                "questions": [{"prompt": "How many looms do you own?", "isRequired": True}],
            }],
        },
        headers=_headers(world),
    )
    assert created.status_code == 201, created.text

    # The full, uncapped list. Same load path as the download, so this is the warning existing at
    # all rather than a second opinion about it.
    preview = client.get(
        f"/api/design-workshops/{workshop_id}/report/preview", headers=_headers(world)
    )
    assert preview.status_code == 200, preview.text
    assert any("questionnaire annexure" in w for w in preview.json()["warnings"]), (
        "an attached questionnaire with no recorded answers raised no warning at all, so its "
        "absence from the report is invisible on every surface"
    )

    generated = client.post(
        f"/api/design-workshops/{workshop_id}/report",
        json={"formats": ["DOCX"], "record": False},
        headers=_headers(world),
    )
    assert generated.status_code == 200, generated.text
    header = generated.headers["x-report-warnings"]
    pieces = [piece for piece in header.split("; ") if piece]

    assert "questionnaire annexure" in header or "further warning(s) did not fit" in header, (
        "the questionnaire annexure's warning did not reach the header AND the header did not say "
        f"anything was dropped, so the designer is told nothing at all. Header was: {header!r}"
    )
    # The count is the TRUE total, never reduced to what fitted — a client comparing the two can
    # tell it is not holding the whole list.
    assert int(generated.headers["x-report-warning-count"]) >= len(pieces)


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_questionnaire_title_a_designer_typed_cannot_split_its_own_warning(client, world):
    """THE TITLE IS DESIGNER-TYPED AND IT REACHES A HEADER WHOSE SEPARATOR IS ";".

    ``questionnaire_warnings`` interpolates the questionnaire's title into the sentence that tells a
    designer their annexure is missing, and ``frontend/lib/designWorkshops.ts`` splits
    ``x-report-warnings`` on ``";"`` and prints each piece as its own warning. So a form called
    "Loom survey; round two" — an ordinary thing to call a second round of fieldwork — delivered
    that one sentence as two, the second of them ``"round two)."``, which says nothing at all on its
    own. Measured against the running server before this was fixed: count 2, three pieces on screen.

    PHOTO_CATALOGUE rather than the default, deliberately: it raises two warnings, so nothing is
    near the header's budget and this pins the SEPARATOR question alone rather than re-testing the
    truncation the test above owns.

    The split asserted here is the frontend's — ``";"``, not ``"; "`` — because what is pinned is
    what the designer reads.
    """
    workshop = client.post(
        "/api/design-workshops",
        json={"title": f"Split title {world['stamp']}", "templateId": "PHOTO_CATALOGUE"},
        headers=_headers(world),
    )
    assert workshop.status_code == 201, workshop.text
    workshop_id = workshop.json()["id"]

    title = f"Loom survey; round two {world['stamp']}"
    created = client.post(
        "/api/questionnaires",
        json={
            "title": title,
            "designWorkshopId": workshop_id,
            "sections": [{
                "title": "About the craft",
                "code": "CRAFT",
                "questions": [{"prompt": "How many looms do you own?", "isRequired": True}],
            }],
        },
        headers=_headers(world),
    )
    assert created.status_code == 201, created.text

    generated = client.post(
        f"/api/design-workshops/{workshop_id}/report",
        json={"formats": ["DOCX"], "record": False},
        headers=_headers(world),
    )
    assert generated.status_code == 200, generated.text
    header = generated.headers["x-report-warnings"]
    count = int(generated.headers["x-report-warning-count"])
    assert "further warning(s) did not fit" not in header, (
        "this fixture is meant to fit inside the budget whole; if it stopped fitting, the "
        "assertion below is no longer measuring the separator"
    )

    pieces = [piece.strip() for piece in header.split(";") if piece.strip()]
    assert len(pieces) == count, (
        f"x-report-warning-count says {count} but the designer's screen shows {len(pieces)} "
        f"warnings, so a designer-typed semicolon split one sentence into halves: {pieces}"
    )
    annexure = [piece for piece in pieces if "questionnaire annexure" in piece]
    assert len(annexure) == 1 and annexure[0].endswith(")."), (
        f"the annexure's own warning did not arrive as one finished sentence: {pieces}"
    )


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_colleague_may_answer_a_questionnaire_but_may_not_reword_it(client, world):
    """The split the access model turns on: handing somebody a form to fill in must not hand them
    the ability to reword it halfway through the fieldwork."""
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Shared {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[{"code": "A", "title": "A", "questions": [{"id": "", "prompt": "How many looms?"}]}],
        ),
    ).json()["questionnaire"]
    question = _questions(created)["How many looms?"]
    colleague = _headers(world, "colleague")

    entry = client.post(
        f"/api/questionnaires/{created['id']}/entries",
        json={"respondentName": "Sita"},
        headers=colleague,
    )
    assert entry.status_code == 201, entry.text
    saved = client.put(
        f"/api/questionnaires/{created['id']}/entries/{entry.json()['id']}/answers",
        json={"answers": [{"questionId": question["id"], "answerText": "4"}]},
        headers=colleague,
    )
    assert saved.status_code == 200, saved.text

    refused = client.patch(
        f"/api/questionnaires/{created['id']}/questions/{question['id']}",
        json={"prompt": "Reworded by somebody else"},
        headers=colleague,
    )
    assert refused.status_code == 403, refused.text


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_designer_cannot_rewrite_whose_sitting_a_colleagues_answers_belong_to(
    client, world
):
    """THE UNGUARDED WRITE.

    ``update_entry`` checked that the caller was A designer and that the sitting belonged to the
    named questionnaire, and nothing else — so any designer could PATCH any sitting and overwrite
    ``respondentName``, which is the NAME OF THE PERSON INTERVIEWED. That is one designer
    relabelling whose testimony a recorded answer belongs to, on a research instrument whose
    report goes to a ministry, with nothing in the document to show it happened. The route needs
    only the entry id, and ids travel in URLs.

    The sitting is recorded by the colleague here, so the person refused is the questionnaire's
    OWNER's peer rather than a stranger: neither the form nor the sitting is theirs.
    """
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Authorship {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[{"code": "A", "title": "A", "questions": [{"id": "", "prompt": "Q?"}]}],
        ),
    ).json()["questionnaire"]
    colleague = _headers(world, "colleague")

    entry = client.post(
        f"/api/questionnaires/{created['id']}/entries",
        json={"respondentName": "Sita Devi"},
        headers=colleague,
    ).json()

    # A third designer — neither the sitting's author nor the form's owner.
    stranger = _headers(world, "stranger")
    refused = client.patch(
        f"/api/questionnaires/{created['id']}/entries/{entry['id']}",
        json={"respondentName": "Somebody Else"},
        headers=stranger,
    )
    assert refused.status_code == 403, refused.text

    # And the name is still the one the interviewer recorded.
    reread = client.get(
        f"/api/questionnaires/{created['id']}", headers=colleague
    ).json()["entries"]
    assert [e["respondentName"] for e in reread] == ["Sita Devi"]


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_the_interviewer_and_the_form_owner_may_both_correct_a_sitting(client, world):
    """The other half, and the reason this is not simply ``_require_owner``: the person who ran
    the sitting mistypes the respondent's name and has to be able to fix it, and the designer
    whose instrument it is has to be able to correct their own fieldwork."""
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Correctable {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[{"code": "A", "title": "A", "questions": [{"id": "", "prompt": "Q?"}]}],
        ),
    ).json()["questionnaire"]
    colleague = _headers(world, "colleague")

    entry = client.post(
        f"/api/questionnaires/{created['id']}/entries",
        json={"respondentName": "Sta Devi"},
        headers=colleague,
    ).json()

    by_author = client.patch(
        f"/api/questionnaires/{created['id']}/entries/{entry['id']}",
        json={"respondentName": "Sita Devi"},
        headers=colleague,
    )
    assert by_author.status_code == 200, by_author.text

    by_owner = client.patch(
        f"/api/questionnaires/{created['id']}/entries/{entry['id']}",
        json={"notes": "Recorded through a translator."},
        headers=_headers(world),
    )
    assert by_owner.status_code == 200, by_owner.text

    by_admin = client.patch(
        f"/api/questionnaires/{created['id']}/entries/{entry['id']}",
        json={"title": "Sitting 1"},
        headers=_headers(world, "admin"),
    )
    assert by_admin.status_code == 200, by_admin.text


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_another_designers_sittings_do_not_ride_along_on_the_form(client, world):
    """Reading the FORM stays open — a colleague handed it has to be able to fill it in — and the
    SITTINGS recorded against it are a different thing wearing the same payload.

    Each one carries the respondent's name, the interviewer's notes and every recorded answer.
    Any designer holding the id got all of it, while their own list view showed ``total: 0`` for
    the same questionnaire: the app said it did not exist and the API handed it over.
    """
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Sittings {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[{"code": "A", "title": "A", "questions": [{"id": "", "prompt": "Q?"}]}],
        ),
    ).json()["questionnaire"]
    question = _questions(created)["Q?"]

    entry = client.post(
        f"/api/questionnaires/{created['id']}/entries",
        json={"respondentName": "Sita Devi", "notes": "Her account of the dye pit."},
        headers=_headers(world),
    ).json()
    client.put(
        f"/api/questionnaires/{created['id']}/entries/{entry['id']}/answers",
        json={"answers": [{"questionId": question["id"], "answerText": "Twelve looms"}]},
        headers=_headers(world),
    )

    stranger = _headers(world, "stranger")
    seen = client.get(f"/api/questionnaires/{created['id']}", headers=stranger)
    assert seen.status_code == 200, "the FORM stays readable, or a colleague cannot fill it in"
    payload = seen.json()
    assert payload["entries"] == [], "another designer's sittings must not ride along"
    body = seen.text
    for secret in ("Sita Devi", "Her account of the dye pit.", "Twelve looms"):
        assert secret not in body, f"{secret!r} reached a designer who recorded none of it"

    # The questions themselves are untouched, and so is the fact that one of them has answers —
    # that is a property of the question, not of who is looking at it.
    assert _questions(payload)["Q?"]["hasAnswers"] is True

    # The owner still sees their own fieldwork in full.
    mine = client.get(f"/api/questionnaires/{created['id']}", headers=_headers(world)).json()
    assert [e["respondentName"] for e in mine["entries"]] == ["Sita Devi"]


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_the_workbook_of_somebody_elses_answers_is_not_a_public_download(client, world):
    """The second door onto the same data, and a wider one: ``export_payload`` is deliberately
    lossless, so the .xlsx carries every sitting, every respondent's name and every answer in one
    file. Narrowing the JSON while leaving this open would have moved the leak, not closed it."""
    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"Workbook {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[{"code": "A", "title": "A", "questions": [{"id": "", "prompt": "Q?"}]}],
        ),
    ).json()["questionnaire"]

    refused = client.get(
        f"/api/questionnaires/{created['id']}/xlsx", headers=_headers(world, "stranger")
    )
    assert refused.status_code == 403, refused.text

    mine = client.get(f"/api/questionnaires/{created['id']}/xlsx", headers=_headers(world))
    assert mine.status_code == 200, mine.text


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_the_database_itself_refuses_to_orphan_an_answer(client, world):
    """THE FLOOR UNDER THE WHOLE RULE.

    Everything above is enforced in Python, and a rule that lives only in Python is one
    ``delete_many`` away from orphaning a fortnight of somebody's fieldwork — written by a future
    contributor who never read the service. So this reaches PAST the service layer and deletes an
    answered question directly through Prisma. Postgres must refuse it.

    Its OWN Prisma client, not the ``db`` singleton the app uses. That one is bound to the
    TestClient's event loop, and awaiting it from this test's loop fails with "bound to a different
    event loop" rather than with anything to do with the constraint under test — which is exactly
    the cross-loop trap the ``world`` fixture's docstring warns about.
    """
    from prisma.errors import ForeignKeyViolationError

    from prisma import Prisma

    created = _upload(
        client,
        world,
        build_questionnaire_workbook(
            title=f"FK floor {world['stamp']}",
            description=None,
            questionnaire_id="",
            version=1,
            sections=[
                {
                    "code": "A",
                    "title": "A",
                    "questions": [{"id": "", "prompt": "Answered", "answers": {"R": "12"}}],
                }
            ],
            entry_labels=["R"],
        ),
    ).json()["questionnaire"]
    question_id = _questions(created)["Answered"]["id"]

    own = Prisma()
    await own.connect()
    try:
        with pytest.raises(ForeignKeyViolationError):
            await own.questionnaireformquestion.delete(where={"id": question_id})
        # Still there, still answered.
        survivor = await own.questionnaireformquestion.find_unique(where={"id": question_id})
        assert survivor is not None
    finally:
        await own.disconnect()


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_question_someone_only_tabbed_past_can_still_be_deleted(client, world):
    """THE BLANK-ANSWER FOREIGN KEY, which reached a designer as a bare 500.

    `guard_question_edit` decides retire-vs-delete on whether an answer has actual TEXT, and a row
    saved as "" or "   " deliberately does not count — it must not freeze a question nobody ever
    answered. That judgement is right. But the ROW still exists and
    `QuestionnaireFormAnswer.questionId` is `onDelete: Restrict`, so Postgres refused the delete and
    the response was "Something went wrong on the server" with a ForeignKeyViolationError in the log
    and nothing on screen a designer could act on.

    It is the ordinary path, not an edge case. The app writes exactly these rows when somebody opens
    a sitting, tabs through it and saves — which is what a designer does the first time they look at
    their own form. Then they tidy up a question they mistyped, and it breaks.
    """
    created = client.post(
        "/api/questionnaires",
        json={"title": f"Tabbed past {world['stamp']}"},
        headers=_headers(world),
    )
    assert created.status_code == 201, created.text
    form = created.json()
    questionnaire_id = form["id"]

    section = client.post(
        f"/api/questionnaires/{questionnaire_id}/sections",
        json={"code": "S1", "title": "Section one"},
        headers=_headers(world),
    )
    assert section.status_code == 201, section.text
    section_id = section.json()["sections"][-1]["id"]

    made = client.post(
        f"/api/questionnaires/{questionnaire_id}/sections/{section_id}/questions",
        json={"prompt": "How many looms?"},
        headers=_headers(world),
    )
    assert made.status_code == 201, made.text
    question_id = _questions(made.json())["How many looms?"]["id"]

    entry = client.post(
        f"/api/questionnaires/{questionnaire_id}/entries",
        json={"title": "A sitting somebody opened and left"},
        headers=_headers(world),
    )
    assert entry.status_code == 201, entry.text
    entry_id = entry.json()["id"]

    # Whitespace, not None: this is what the form sends for a box that was focused and left alone.
    saved = client.put(
        f"/api/questionnaires/{questionnaire_id}/entries/{entry_id}/answers",
        json={"answers": [{"questionId": question_id, "answerText": "   "}]},
        headers=_headers(world),
    )
    assert saved.status_code == 200, saved.text

    removed = client.delete(
        f"/api/questionnaires/{questionnaire_id}/questions/{question_id}",
        headers=_headers(world),
    )
    assert removed.status_code == 200, removed.text
    assert removed.json()["action"] == "deleted", (
        "a question nobody actually answered must DELETE, not retire — retiring it would leave the "
        "designer's mistyped question in the form for ever"
    )
    assert not _questions(removed.json()["questionnaire"]), "the question is gone from the form"


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_question_with_a_real_answer_is_still_retired_rather_than_deleted(client, world):
    """The other half of the same rule, pinned so the fix above cannot be widened into data loss."""
    created = client.post(
        "/api/questionnaires",
        json={"title": f"Really answered {world['stamp']}"},
        headers=_headers(world),
    )
    questionnaire_id = created.json()["id"]
    section = client.post(
        f"/api/questionnaires/{questionnaire_id}/sections",
        json={"code": "S1", "title": "Section one"},
        headers=_headers(world),
    )
    section_id = section.json()["sections"][-1]["id"]
    made = client.post(
        f"/api/questionnaires/{questionnaire_id}/sections/{section_id}/questions",
        json={"prompt": "How many looms?"},
        headers=_headers(world),
    )
    question_id = _questions(made.json())["How many looms?"]["id"]
    entry = client.post(
        f"/api/questionnaires/{questionnaire_id}/entries",
        json={"title": "A real sitting"},
        headers=_headers(world),
    )
    client.put(
        f"/api/questionnaires/{questionnaire_id}/entries/{entry.json()['id']}/answers",
        json={"answers": [{"questionId": question_id, "answerText": "Twelve"}]},
        headers=_headers(world),
    )

    removed = client.delete(
        f"/api/questionnaires/{questionnaire_id}/questions/{question_id}",
        headers=_headers(world),
    )
    assert removed.status_code == 200, removed.text
    assert removed.json()["action"] == "retired"
    kept = _questions(removed.json()["questionnaire"])["How many looms?"]
    assert kept["isActive"] is False
