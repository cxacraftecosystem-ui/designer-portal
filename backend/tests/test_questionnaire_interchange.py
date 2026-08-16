"""QUESTIONNAIRE INTERCHANGE: sending a question set to another designer, and receiving one.

This file exists because of three holes that were found together, all of them in the seam between
"my questionnaire" and "yours". They are separate defects with one cause: there was exactly ONE
artefact — a lossless workbook carrying every sitting, every respondent's name and every answer — and
it was being asked to do two incompatible jobs at once.

**HOLE 1 — THERE WAS NO QUESTIONS-ONLY EXPORT, SO SHARING WAS IMPOSSIBLE.** ``GET
/questionnaires/{id}/xlsx`` refuses anyone who is not the owner, a designer on its design workshop,
or an admin, because the file it serves really is somebody's fieldwork. That refusal is correct. The
consequence was that a designer who simply wanted to give a colleague their eighteen questions had
no way to do it, and the only apparent fix — loosening that gate — would have moved a leak rather
than closing one. The fix is a SECOND artefact: ``/question-set.xlsx``, questions and nothing else,
which any designer may take. The tests below pin the property that makes it safe to hand over: it
carries no answers and no respondent names, anywhere in the file.

**HOLE 2 — THE IMPORT SILENTLY STOLE THE SENDER'S FIELDWORK.** ``create_from_parsed`` ignored
``parsed.questionnaireId`` and wrote the workbook's answers as sittings with ``answeredById`` and
``createdById`` set to THE UPLOADER. So designer B, handed designer A's workbook, uploaded it and
acquired A's respondents' answers as B's own recorded fieldwork, under B's name — and
``QUESTIONNAIRE_ANNEXURE`` is in all six report templates, so it went into the .docx B submits to a
ministry. Nothing on any screen said so. Two tests below are that exact sequence.

**HOLE 3 — THE DEFAULT QUESTIONNAIRE DID NOT HONOUR THE SUPERSEDE RULE.** The custom builder has
always honoured it; ``PATCH /questionnaire/questions/{id}`` — on the ONE global instrument every
researcher answers — wrote the new prompt straight over the old one with no answer check at all, so
every answer already given silently re-attributed to the new wording. The last section of this file
is the same test the custom builder already has, aimed at the other feature.

WHAT NEEDS WHAT. The parser and writer tests need only openpyxl. Everything else needs Postgres,
because the rules are statements about rows; that half skips itself when ``DATABASE_URL`` is not
local, exactly as ``test_questionnaire_forms`` does.

    docker compose up -d postgres
    cd backend && .venv/Scripts/python.exe -m prisma migrate deploy --schema prisma/schema.prisma
"""

import os
import uuid
import zipfile
from io import BytesIO
from typing import Any

import pytest

from app.core.db import db
from app.core.security import create_access_token
from app.services.questionnaire_xlsx import (
    QUESTION_SET_CONTENTS,
    build_pro_forma,
    build_question_set_workbook,
    build_questionnaire_workbook,
    download_filename,
    parse_questionnaire_workbook,
    question_set_filename,
)

_URL = os.environ.get("DATABASE_URL", "")
_LOCAL = any(host in _URL for host in ("localhost", "127.0.0.1"))

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# The instrument under test everywhere below: two sections, and answers recorded against both — the
# only state in which "questions only" and "everything" are different files.
ANSWERED_FORM = [
    {
        "code": "CRAFT",
        "title": "About the craft",
        "questions": [
            {
                "id": "q-looms",
                "prompt": "How many looms do you own?",
                "helpText": "Count working looms only",
                "isRequired": True,
                "answers": {"Ramesh Devi": "12"},
                "answerNotes": {"Ramesh Devi": "answered through a translator"},
            },
            {"id": "q-taught", "prompt": "Who taught you?", "isRequired": False, "answers": {}},
        ],
    },
    {
        "code": "MAT",
        "title": "Materials",
        "questions": [
            {
                "id": "q-yarn",
                "prompt": "Where do you buy your yarn?",
                "isRequired": True,
                "answers": {"Ramesh Devi": "Panipat mandi"},
            }
        ],
    },
]


def _every_string_in(payload: bytes) -> str:
    """Every byte of text in the workbook, from every part of the .xlsx archive.

    NOT a substring search over the raw bytes — an .xlsx is a ZIP, so its text is DEFLATE-compressed
    and a respondent's name that is plainly present would not be found by ``b"Ramesh" in payload``.
    A test that searched the compressed bytes would pass on a file that leaks, which is the one
    outcome worse than no test at all. Cell text lives in ``xl/sharedStrings.xml`` rather than in the
    sheet, so the whole archive is read.
    """
    with zipfile.ZipFile(BytesIO(payload)) as archive:
        return "\n".join(
            archive.read(name).decode("utf-8", "replace") for name in archive.namelist()
        )


def _flatten(parsed: Any) -> list[tuple[str, str]]:
    return [
        (section.title, question.prompt)
        for section in parsed.sections
        for question in section.questions
    ]


# --------------------------------------------------------------------------------------
# 1. HOLE 1 — the questions-only workbook, as a file. No database needed.
# --------------------------------------------------------------------------------------


def _question_set_of(sections: list[dict[str, Any]]) -> bytes:
    """The same sections a full workbook would carry, exported as a QUESTION SET.

    Answers are stripped and ids blanked here rather than by the writer, because that is what
    ``export_question_set_payload`` does in the service: the writer is handed clean data and does not
    have to be trusted to filter. Building the fixture the same way is what keeps this test honest
    about which layer it is pinning.
    """
    return build_question_set_workbook(
        title="Loom study",
        description="Weaving cluster survey",
        sections=[
            {
                "code": section["code"],
                "title": section["title"],
                "questions": [
                    {
                        "id": "",
                        "prompt": q["prompt"],
                        "helpText": q.get("helpText"),
                        "isRequired": q.get("isRequired", False),
                        "answers": {},
                        "answerNotes": {},
                    }
                    for q in section["questions"]
                ],
            }
            for section in sections
        ],
        source_title="Loom study",
        shared_by="A Designer",
        exported_on="2026-08-16",
    )


def test_a_question_set_carries_no_answers_and_no_respondent_names_anywhere_in_the_file():
    """THE ONE PROPERTY THAT MAKES THIS ARTEFACT SAFE TO SEND, checked over the whole archive.

    A question set is offered to any designer precisely because it contains no fieldwork. If a
    respondent's name or an answer can survive into it — through a stray column, a cached value, a
    future "helpful" addition of a preview row — then the wider gate on it becomes a leak, and the
    leak would be invisible: the file still looks like a list of questions.

    So this does not merely check that the parser reads no answers back. It reads every part of the
    .xlsx and asserts the strings are not in the file AT ALL.
    """
    payload = _question_set_of(ANSWERED_FORM)
    text = _every_string_in(payload)

    assert "Ramesh Devi" not in text, "a respondent's name reached a file advertised as carrying none"
    assert "Panipat mandi" not in text, "a recorded answer reached the questions-only export"
    assert "answered through a translator" not in text, "an interviewer's note on an answer got out"

    # And the questions themselves did survive, or the file is safe and useless.
    assert "How many looms do you own?" in text
    assert "Count working looms only" in text


def test_a_question_set_round_trips_as_a_brand_new_questionnaire_rather_than_an_edit():
    """Blank Question IDs and a blank Questionnaire ID, and both blanks are load-bearing.

    A Question ID belongs to the SENDER's database. Left in, the receiving designer's re-upload
    reports every single row as "Question ID … does not belong to this questionnaire", and — the part
    that matters — those ids are one of the two signals the import uses to recognise a workbook that
    came out of the app and therefore to refuse to re-record its answers. A shared question set has
    to read as exactly what it is: a filled-in pro-forma.

    A Questionnaire ID on the Details sheet is the same problem from the other side: with one,
    ``POST /questionnaires/{id}/upload`` answers 409 "that workbook was downloaded from a different
    questionnaire" — right for the full workbook, wrong for a question set the receiver may use.
    """
    parsed = parse_questionnaire_workbook(_question_set_of(ANSWERED_FORM), filename="shared.xlsx")

    assert parsed.questionnaireId is None, "the question set must not claim to be an edit of its source"
    assert [q.questionId for s in parsed.sections for q in s.questions] == [None, None, None]
    assert parsed.entryLabels == [], "an answer column with a sitting's name on it came through"
    assert not parsed.hasAnswers
    assert _flatten(parsed) == [
        ("About the craft", "How many looms do you own?"),
        ("About the craft", "Who taught you?"),
        ("Materials", "Where do you buy your yarn?"),
    ]
    # The required flags and help text travel too — a question set that lost them would be a list of
    # sentences rather than an instrument.
    required = {q.prompt: q.isRequired for s in parsed.sections for q in s.questions}
    assert required == {
        "How many looms do you own?": True,
        "Who taught you?": False,
        "Where do you buy your yarn?": True,
    }


def test_the_provenance_rows_on_the_details_sheet_do_not_overwrite_the_description():
    """A LABEL COLLISION HERE WOULD CORRUPT THE RECEIVER'S QUESTIONNAIRE, silently.

    ``_read_details`` matches column A against a family of aliases, and the aliases for
    ``description`` include "notes", "summary", "about" and "purpose". The question set adds
    provenance rows — Contents, Exported from, Shared by, Exported on — and had any of them been
    spelled "Notes" or "About", the sentence written for a human reader would come back in as the
    questionnaire's DESCRIPTION on import: the receiver's form silently described as "Questions only
    — no answers…". This pins the labels apart.
    """
    parsed = parse_questionnaire_workbook(_question_set_of(ANSWERED_FORM), filename="shared.xlsx")

    assert parsed.title == "Loom study"
    assert parsed.description == "Weaving cluster survey"

    # The provenance really is in the file — it is only kept out of the parsed fields.
    text = _every_string_in(_question_set_of(ANSWERED_FORM))
    assert QUESTION_SET_CONTENTS in text
    assert "A Designer" in text
    assert "2026-08-16" in text


def test_the_two_downloads_have_different_file_names():
    """Both files land in the same Downloads folder with the same questionnaire title on them.

    One of them is a list of questions and the other is every respondent that designer has ever
    interviewed. The name is the last thing standing between a colleague and the wrong attachment.
    """
    assert download_filename("Loom study") == "Loom-study.xlsx"
    assert question_set_filename("Loom study") == "Loom-study-questions.xlsx"
    assert question_set_filename(None) == "questionnaire-questions.xlsx"


def test_the_pro_forma_still_parses_after_the_details_sheet_learned_extra_rows():
    """The note paragraph moved from a fixed row 8 to "below whatever rows exist".

    With the four base pairs it still lands on row 8 — but if that arithmetic is ever wrong the note
    lands ON a value row and eats it, which for the blank pro-forma would mean the Questionnaire ID
    row carrying a paragraph of prose that ``_read_details`` would then hand back as an id.
    """
    from openpyxl import load_workbook

    book = load_workbook(BytesIO(build_pro_forma()))
    details = book["Details"]
    assert [details.cell(row=row, column=1).value for row in range(3, 7)] == [
        "Questionnaire title",
        "Description",
        "Questionnaire ID",
        "Version",
    ]
    assert details.cell(row=5, column=2).value in (None, ""), (
        "the blank pro-forma must carry no Questionnaire ID, or every upload of it would read as an "
        "edit of some other questionnaire"
    )


# --------------------------------------------------------------------------------------
# 2. Everything below needs Postgres.
# --------------------------------------------------------------------------------------


@pytest.fixture(scope="module")
def anyio_backend():
    return "asyncio"


@pytest.fixture(scope="module")
async def world():
    """A designer who owns a questionnaire, a second designer who is sent one, and an admin.

    Rows are created here rather than inside a test for the reason ``test_questionnaire_forms``
    states: the Prisma client is shared with the running app and bound to the TestClient's event
    loop, so touching it from a test's own loop is the kind of cross-loop use that fails
    intermittently rather than honestly.

    The ADMIN doubles as the questionnaire MANAGER for the section-3 tests: ``can_manage_questionnaire``
    is Professor-and-above, and ADMIN outranks Professor. A designer deliberately does NOT manage the
    global questionnaire — that gate is unchanged by this work and is not this file's subject.
    """
    if not _LOCAL:
        pytest.skip("needs a local database")
    from fastapi.testclient import TestClient

    from app.main import app

    stamp = uuid.uuid4().hex[:8]
    people: dict[str, Any] = {}
    await db.connect()
    try:
        for slug, role, name in (
            ("owner", "DESIGNER", "Questionnaire Owner"),
            ("receiver", "DESIGNER", "Receiving Designer"),
            ("admin", "ADMIN", "Interchange Admin"),
        ):
            people[slug] = await db.user.create(
                data={
                    "email": f"qi-{slug}-{stamp}@example.org",
                    "name": name,
                    "role": role,
                }
            )
    finally:
        await db.disconnect()
    with TestClient(app) as client:
        yield {"client": client, "people": people, "stamp": stamp}


@pytest.fixture
def client(world):
    return world["client"]


def _headers(world: dict[str, Any], slug: str = "owner") -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token(subject=world['people'][slug].id)}"}


def _upload(client: Any, world: dict[str, Any], payload: bytes, slug: str = "owner", name: str = "form.xlsx") -> Any:
    return client.post(
        "/api/questionnaires/upload",
        files={"file": (name, payload, XLSX_MIME)},
        headers=_headers(world, slug),
    )


def _questions(form: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {q["prompt"]: q for section in form["sections"] for q in section["questions"]}


def _hand_typed_workbook(title: str, *, with_answers: bool) -> bytes:
    """A workbook as a designer builds it: no Questionnaire ID, no Question IDs, maybe answers.

    This is the ORDINARY path — the designer who ran interviews on paper and typed them up — and it
    must keep working exactly as it always has. It is built through ``build_questionnaire_workbook``
    with an empty ``questionnaire_id`` and empty per-question ids, which is byte-for-byte the shape a
    filled-in pro-forma has.
    """
    answers = {"Ramesh Devi": "12"} if with_answers else {}
    return build_questionnaire_workbook(
        title=title,
        description=None,
        questionnaire_id="",
        version=1,
        sections=[
            {
                "code": "CRAFT",
                "title": "About the craft",
                "questions": [
                    {"id": "", "prompt": "How many looms do you own?", "answers": answers},
                    {"id": "", "prompt": "Who taught you?", "answers": {}},
                ],
            }
        ],
        entry_labels=["Ramesh Devi"] if with_answers else [],
    )


# --------------------------------------------------------------------------------------
# 3. HOLE 1, through the API — who may take which file
# --------------------------------------------------------------------------------------


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_the_question_set_is_open_to_any_designer_while_the_full_workbook_is_not(client, world):
    """THE WHOLE POINT OF THERE BEING TWO FILES, asserted as one pair of responses.

    Before this, a designer who wanted a colleague's questions had exactly one endpoint to ask, and
    it answered 403 — correctly, because that file carries every sitting recorded against the
    questionnaire. Sharing was therefore impossible, and the tempting fix was to widen that gate,
    which would have handed over the respondents along with the questions.

    So: the SAME questionnaire, the SAME caller, two endpoints, two answers. If a future change makes
    these two agree, one of the two features has been destroyed — either sharing is impossible again,
    or somebody's fieldwork just became a public download.
    """
    created = _upload(client, world, _hand_typed_workbook(f"Shared {world['stamp']}", with_answers=True))
    assert created.status_code == 201, created.text
    questionnaire_id = created.json()["questionnaire"]["id"]

    refused = client.get(
        f"/api/questionnaires/{questionnaire_id}/xlsx", headers=_headers(world, "receiver")
    )
    assert refused.status_code == 403, refused.text
    # The refusal names the next move rather than merely saying no — that is the house rule, and here
    # it is also the only way the receiving designer learns the other file exists.
    assert "question set" in refused.json()["detail"]

    allowed = client.get(
        f"/api/questionnaires/{questionnaire_id}/question-set.xlsx",
        headers=_headers(world, "receiver"),
    )
    assert allowed.status_code == 200, allowed.text
    assert allowed.headers["content-type"].startswith(XLSX_MIME)
    assert "-questions.xlsx" in allowed.headers["content-disposition"]


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_the_question_set_served_by_the_api_carries_no_recorded_answer(client, world):
    """The property from section 1, re-checked on the bytes the SERVER produced.

    Section 1 pins the writer. This pins the whole chain a designer actually presses —
    ``load_question_set`` -> ``export_question_set_payload`` -> ``build_question_set_workbook`` — on a
    questionnaire that genuinely has a sitting recorded against it in the database. A filter that was
    correct in the writer and forgotten in the loader would pass every test in section 1.
    """
    created = _upload(client, world, _hand_typed_workbook(f"Served {world['stamp']}", with_answers=True))
    assert created.status_code == 201, created.text
    form = created.json()["questionnaire"]
    assert len(form["entries"]) == 1, "the fixture must actually have a sitting, or this proves nothing"

    served = client.get(
        f"/api/questionnaires/{form['id']}/question-set.xlsx", headers=_headers(world, "receiver")
    )
    assert served.status_code == 200, served.text
    text = _every_string_in(served.content)
    assert "Ramesh Devi" not in text
    assert "How many looms do you own?" in text


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_retired_question_is_not_sent_on_in_a_question_set(client, world):
    """A retired question is kept because ANSWERS hang off it, not because it is still asked.

    Sending one would plant a question the sender deliberately replaced into the receiver's
    brand-new, empty form — sitting next to the replacement, with nothing to say which of the two is
    the real one. The full workbook DOES carry retired questions, and must, because its round trip
    has to be lossless; this is the one place the two exports differ for a reason other than privacy.
    """
    created = _upload(client, world, _hand_typed_workbook(f"Retired {world['stamp']}", with_answers=True))
    form = created.json()["questionnaire"]
    looms = _questions(form)["How many looms do you own?"]
    assert looms["hasAnswers"] is True

    superseded = client.patch(
        f"/api/questionnaires/{form['id']}/questions/{looms['id']}",
        json={"prompt": "How many weavers work with you?"},
        headers=_headers(world),
    )
    assert superseded.status_code == 200, superseded.text
    assert superseded.json()["action"] == "superseded"

    served = client.get(
        f"/api/questionnaires/{form['id']}/question-set.xlsx", headers=_headers(world)
    )
    text = _every_string_in(served.content)
    assert "How many weavers work with you?" in text, "the wording being asked today must travel"
    assert "How many looms do you own?" not in text, (
        "the retired wording was sent on; it is kept only because answers hang off it and it is not "
        "part of the instrument any more"
    )


# --------------------------------------------------------------------------------------
# 4. HOLE 2 — the import that used to steal the sender's fieldwork
# --------------------------------------------------------------------------------------


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_importing_a_question_set_creates_an_empty_questionnaire(client, world):
    """The receiving half of the sharing loop: the questions arrive, and nothing else does.

    An empty questionnaire is the CORRECT outcome and not a disappointing one — the receiver is
    about to run their own fieldwork with this instrument. If any sitting appeared here it would be
    the sender's, wearing the receiver's name.
    """
    sent = _upload(client, world, _hand_typed_workbook(f"To send {world['stamp']}", with_answers=True))
    source_id = sent.json()["questionnaire"]["id"]

    question_set = client.get(
        f"/api/questionnaires/{source_id}/question-set.xlsx", headers=_headers(world, "receiver")
    )
    assert question_set.status_code == 200, question_set.text

    received = _upload(client, world, question_set.content, slug="receiver", name="from-a-colleague.xlsx")
    assert received.status_code == 201, received.text
    body = received.json()
    form = body["questionnaire"]

    assert form["questionCount"] == 2
    assert _questions(form).keys() == {"How many looms do you own?", "Who taught you?"}
    assert form["entries"] == [], "the sender's sitting arrived in the receiver's questionnaire"
    assert body["report"]["entriesCreated"] == 0
    assert body["report"]["answersImported"] == 0
    assert body["report"]["answersSkipped"] == 0
    # Nothing to explain, because nothing was refused: a question set has no answers to decide about.
    assert body["report"]["provenance"] is None
    assert form["ownerId"] == world["people"]["receiver"].id

    # And the sender still has their own sitting. A "share" that moved it would be worse than one
    # that copied it.
    original = client.get(f"/api/questionnaires/{source_id}", headers=_headers(world)).json()
    assert len(original["entries"]) == 1


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_uploading_someone_elses_full_workbook_never_records_their_answers_as_yours(
    client, world
):
    """THE DATA-INTEGRITY BUG, reproduced end to end and then pinned shut.

    THE OLD BEHAVIOUR, exactly: designer B obtains designer A's downloaded workbook — from an email,
    a shared drive, a handover — and uploads it. ``create_from_parsed`` ignored the Questionnaire ID
    on its Details sheet and wrote every answer column as a sitting with ``createdById`` and
    ``answeredById`` set to B. A's respondent, A's answers, A's interview, now recorded as B's own
    fieldwork under B's name, and printed into the questionnaire annexure of the report B submits.
    Nobody edited anything. No screen mentioned it.

    THE RULE NOW is about the DATA rather than about who is holding the file: a workbook that came
    out of the platform imports its QUESTIONS ONLY, because its answers are already recorded here
    with their true authors on them, and writing them again would duplicate the fieldwork rather than
    import it. So this test does not assert a 403 — B is not being accused of anything, and B does
    get the questions, which is what B legitimately wanted.

    WHAT IT DOES ASSERT is that not one answer row was created for B, and that B is TOLD, in a
    sentence written to be shown as-is. Silence was the whole defect.
    """
    owner_upload = _upload(
        client, world, _hand_typed_workbook(f"Owned {world['stamp']}", with_answers=True)
    )
    assert owner_upload.status_code == 201, owner_upload.text
    owners_form = owner_upload.json()["questionnaire"]
    assert len(owners_form["entries"]) == 1
    assert owners_form["entries"][0]["respondentName"] == "Ramesh Devi"

    # The lossless workbook, as only the owner may take it. This is the file that leaks in the wild.
    full = client.get(f"/api/questionnaires/{owners_form['id']}/xlsx", headers=_headers(world))
    assert full.status_code == 200, full.text
    assert "Ramesh Devi" in _every_string_in(full.content), (
        "the fixture must genuinely carry the respondent, or the theft this test is about cannot "
        "happen and the assertions below are vacuous"
    )

    stolen = _upload(client, world, full.content, slug="receiver", name="someone-elses.xlsx")
    assert stolen.status_code == 201, stolen.text
    body = stolen.json()
    form = body["questionnaire"]

    assert form["entries"] == [], (
        "the sender's sitting was recorded against the uploader's brand-new questionnaire — this is "
        "the defect: somebody else's respondent, answers and interview, now attributed to whoever "
        "opened the file"
    )
    assert body["report"]["entriesCreated"] == 0
    assert body["report"]["answersImported"] == 0
    assert body["report"]["answersSkipped"] >= 1

    provenance = body["report"]["provenance"]
    assert provenance is not None, "the uploader must be told; silence is what made this invisible"
    assert provenance["action"] == "answersNotImported"
    assert provenance["sourceQuestionnaireId"] == owners_form["id"]
    assert "already recorded" in provenance["reason"]
    # Also pushed into `problems`, so every client that already renders the parser's problem list
    # says this without having to be changed first.
    assert any(p["reason"] == provenance["reason"] for p in body["report"]["problems"])

    # The questions DID arrive — the receiver is not punished for the sender's file.
    assert _questions(form).keys() == {"How many looms do you own?", "Who taught you?"}

    # And the owner's own fieldwork is untouched on their own questionnaire.
    still = client.get(f"/api/questionnaires/{owners_form['id']}", headers=_headers(world)).json()
    assert len(still["entries"]) == 1
    assert still["entries"][0]["createdById"] == world["people"]["owner"].id


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_answers_typed_into_a_blank_pro_forma_are_still_imported_and_still_the_uploaders(
    client, world
):
    """THE OTHER HALF OF THE FIX, and the reason it is not simply "refuse workbooks with answers".

    The documented, intended, ordinary case is a designer who ran the interview on paper and typed it
    up in Excel. There is no other recorder anywhere in that picture, so attributing those answers to
    the person who typed them is honest — and a rule that refused them would break the feature for
    the people it was built for. A test that only pinned the refusal would leave a future contributor
    free to "simplify" the check into "any workbook with answers is refused" and nothing would catch
    it.

    The distinguishing fact is in the FILE, not in the account: no Questionnaire ID and no Question
    IDs means it was filled in by hand rather than downloaded out of the platform.
    """
    typed = _upload(
        client,
        world,
        _hand_typed_workbook(f"Typed up {world['stamp']}", with_answers=True),
        slug="receiver",
        name="paper-interviews.xlsx",
    )
    assert typed.status_code == 201, typed.text
    body = typed.json()
    form = body["questionnaire"]

    assert len(form["entries"]) == 1
    entry = form["entries"][0]
    assert entry["source"] == "UPLOAD"
    assert entry["respondentName"] == "Ramesh Devi"
    assert entry["createdById"] == world["people"]["receiver"].id
    assert body["report"]["entriesCreated"] == 1
    assert body["report"]["answersImported"] == 1
    assert body["report"]["answersSkipped"] == 0

    # THE VISIBLE PROVENANCE NOTE. `source = "UPLOAD"` records the same fact in a column no report
    # prints; this sentence prints in the annexure beside the answers, which is the only place a
    # reader can tell "I interviewed this person" from "I typed up a spreadsheet".
    assert entry["notes"] and "uploaded spreadsheet" in entry["notes"]
    assert "paper-interviews.xlsx" in entry["notes"]

    provenance = body["report"]["provenance"]
    assert provenance["action"] == "answersImported"
    assert "attributed to you" in provenance["reason"]


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_workbook_whose_details_sheet_was_deleted_still_gives_its_answers_away_to_nobody(
    client, world
):
    """THE SECOND SIGNAL, and why one was not enough.

    Reading only the Questionnaire ID would have made the whole rule defeatable by clearing a single
    cell — or, with no malice at all, by copying the Questionnaire sheet into a fresh file, which
    loses the Details sheet entirely and is exactly what a designer does when a workbook misbehaves.
    The Question ID column is written by the same export and survives both, so it is asked about too.

    Here the workbook carries question ids and answers and NO questionnaire id: the shape a copied
    sheet has. It must still refuse to re-record the answers.
    """
    copied_sheet = build_questionnaire_workbook(
        title=f"Copied sheet {world['stamp']}",
        description=None,
        # The Details sheet is present but its id cell is empty — the same state as no Details sheet
        # at all, as far as the parser is concerned.
        questionnaire_id="",
        version=1,
        sections=[
            {
                "code": "CRAFT",
                "title": "About the craft",
                "questions": [
                    {
                        "id": "ckzz0000somebodyelses0001",
                        "prompt": "How many looms do you own?",
                        "answers": {"Ramesh Devi": "12"},
                    }
                ],
            }
        ],
        entry_labels=["Ramesh Devi"],
    )
    received = _upload(client, world, copied_sheet, slug="receiver", name="copied.xlsx")
    assert received.status_code == 201, received.text
    body = received.json()

    assert body["questionnaire"]["entries"] == []
    assert body["report"]["answersSkipped"] == 1
    provenance = body["report"]["provenance"]
    assert provenance["action"] == "answersNotImported"
    assert provenance["sourceQuestionnaireId"] is None
    assert "Question ID column" in provenance["reason"]


# --------------------------------------------------------------------------------------
# 5. HOLE 3 — the supersede rule, on the DEFAULT questionnaire
# --------------------------------------------------------------------------------------


def _supersede_columns_generated() -> bool:
    """Whether the generated Prisma client carries ``QuestionnaireQuestion.retiredAt`` and
    ``.supersededById``.

    ================================================================================================
    THIS USED TO BE A SKIP GUARD. IT IS NOW ASSERTED, AND THAT CHANGE IS THE WHOLE POINT.
    ================================================================================================

    The supersede fix needs those two columns. The migration that adds them and the route that reads
    and writes them were written by an agent that — correctly — could not edit ``schema.prisma``,
    because four lanes editing one schema concurrently corrupts the generated client for everybody.
    So the field declarations were handed to the orchestrator, and these five tests carried
    ``@pytest.mark.skipif`` until they landed.

    **They did not land.** The migration went in, the route went in, the model block was lost in the
    concurrent edit — and the suite reported ``12 passed, 5 skipped`` while three live endpoints
    (reword an answered question, delete one, delete a section holding one) answered a bare 500. The
    skip was honest about its reason and still produced the worst possible outcome: the ONLY tests
    that could have caught the breakage were the ones the breakage switched off.

    That is the general rule this note exists to record: **a guard that disables a test exactly when
    the thing it tests is broken is not a guard, it is a blindfold.** The environment guards
    elsewhere in this file (``_LOCAL``, and the rembg/Pillow ones in other modules) are a different
    shape — they skip when the DEVELOPER'S MACHINE cannot answer, not when the PRODUCT is wrong.

    So the introspection stays and the skip goes: the five tests below now run unconditionally, and
    the state they used to tiptoe around is a failing test of its own.
    """
    try:
        from prisma.models import QuestionnaireQuestion
    except (ImportError, RuntimeError):  # pragma: no cover - a client that was never generated
        # prisma-client-py raises RuntimeError ("the client has not been generated yet") rather than
        # ImportError when the package is installed but `prisma generate` has never run.
        return False
    fields = getattr(QuestionnaireQuestion, "model_fields", None)
    if fields is None:  # pydantic v1 clients
        fields = getattr(QuestionnaireQuestion, "__fields__", {})
    return "supersededById" in fields and "retiredAt" in fields


def test_the_supersede_columns_are_in_the_schema_and_not_only_in_the_migration():
    """The model block, the migration and the route must all agree that these two columns exist.

    A migration adds them to the DATABASE; ``schema.prisma`` is what puts them in the generated
    CLIENT; the route reads and writes them through that client. Land two of the three and the
    endpoints 500 — which is exactly what happened, and what this test now catches in one second
    instead of in production.
    """
    assert _supersede_columns_generated(), (
        "QuestionnaireQuestion.retiredAt / .supersededById are missing from the generated Prisma "
        "client. The migration 20260816200000_questionnaire_question_supersede adds them to the "
        "database and app/api/routes/questionnaire.py reads and writes them, so without the model "
        "block in schema.prisma three endpoints answer a bare 500: rewording an answered question, "
        "deleting one, and deleting a section that holds one. Add the two fields to model "
        "QuestionnaireQuestion and run `python -m prisma generate`."
    )



# The four questions of the throwaway section built below, and whether an answer is recorded against
# each. EVERY TEST IN SECTION 5 GETS ITS OWN SUBJECT, deliberately: a shared one makes the tests
# order-dependent, and an order-dependent supersede test is one that passes for the wrong reason the
# moment somebody runs it with `-k`.
_DEFAULT_QUESTIONS = {
    "How many looms do you own?": "12",          # superseded by the first test
    "Who taught you?": None,                     # never answered: the plain-reword and no-stamp case
    "What does a kilo of yarn cost?": "Rs 240",  # answered, left alone: reposition and retire
    "Where do you buy your yarn?": "Panipat",     # answered: the double-supersede refusal, in full
}


@pytest.fixture(scope="module")
async def default_questionnaire(world):
    """One throwaway section of the GLOBAL questionnaire, with answers recorded against three of it.

    THE ANSWERS ARE WRITTEN THROUGH A PRIVATE PRISMA CLIENT rather than through
    ``POST /questionnaire/interviews``, and that is deliberate. Creating an interview by HTTP drags in
    the location requirement, the workshop-submission guard, the artisan-set dedupe and the review
    status policy — four surfaces these tests have no opinion about, any of which failing would look
    like the supersede rule failing. What the rule is about is one row in ``QuestionnaireResponse``,
    so rows are what get created.

    Its OWN client, not the ``db`` singleton: that one is bound to the TestClient's event loop, and
    awaiting it from a test's loop fails with "bound to a different event loop" rather than with
    anything to do with what is under test.

    THE SECTION IS DEACTIVATED ON THE WAY OUT. Unlike a custom questionnaire, which belongs to the
    designer who made it, a section of the global questionnaire appears on every researcher's
    questionnaire screen and as a column of the completion matrix. Leaving test sections switched on
    in a development database is a slow way to make that screen useless.
    """
    if not _LOCAL:
        pytest.skip("needs a local database")
    from prisma import Prisma

    client = world["client"]
    stamp = world["stamp"]
    headers = _headers(world, "admin")

    section = client.post(
        "/api/questionnaire/sections",
        json={"code": f"ZQI{stamp[:5].upper()}", "title": f"Interchange rule {stamp}"},
        headers=headers,
    )
    assert section.status_code == 201, section.text
    section_id = section.json()["id"]

    made: dict[str, str] = {}
    for prompt in _DEFAULT_QUESTIONS:
        question = client.post(
            "/api/questionnaire/questions",
            json={"sectionId": section_id, "prompt": prompt},
            headers=headers,
        )
        assert question.status_code == 201, question.text
        made[prompt] = question.json()["id"]

    own = Prisma()
    await own.connect()
    try:
        interview = await own.questionnaireinterview.create(
            data={
                "title": f"Interchange interview {stamp}",
                "createdById": world["people"]["admin"].id,
            }
        )
        for prompt, answer in _DEFAULT_QUESTIONS.items():
            if answer is None:
                continue
            await own.questionnaireresponse.create(
                data={
                    "interviewId": interview.id,
                    "questionId": made[prompt],
                    "answerText": answer,
                    "answeredById": world["people"]["admin"].id,
                }
            )
    finally:
        await own.disconnect()

    yield {"sectionId": section_id, "questions": made, "interviewId": interview.id}

    cleanup = Prisma()
    await cleanup.connect()
    try:
        await cleanup.questionnairesection.update(
            where={"id": section_id}, data={"isActive": False}
        )
        await cleanup.questionnairequestion.update_many(
            where={"sectionId": section_id}, data={"isActive": False}
        )
    finally:
        await cleanup.disconnect()


def _global_question(client: Any, world: dict[str, Any], question_id: str) -> dict[str, Any] | None:
    """One question of the global questionnaire, retired ones included."""
    sections = client.get(
        "/api/questionnaire/sections?activeOnly=false", headers=_headers(world, "admin")
    )
    assert sections.status_code == 200, sections.text
    for section in sections.json():
        for question in section["questions"]:
            if question["id"] == question_id:
                return question
    return None


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_rewording_an_answered_default_question_supersedes_it_instead_of_overwriting_it(
    client, world, default_questionnaire
):
    """THE DEFECT, ON THE GLOBAL QUESTIONNAIRE: twelve looms became twelve weavers.

    "How many looms do you own?" is answered "12". A professor rewords it to "How many weavers work
    with you?" and — before this fix — one UPDATE statement later, the repository asserted that this
    artisan works with twelve weavers. Nobody edited an answer. Nothing in any log said anything had
    happened. The consolidated per-artisan document, its CSV and the completion matrix all read that
    row, and the .docx built from them goes to a ministry.

    The custom questionnaire builder has refused to do this since it was written. This is the same
    assertion, aimed at the feature one character away in the URL.
    """
    original_id = default_questionnaire["questions"]["How many looms do you own?"]

    reworded = client.patch(
        f"/api/questionnaire/questions/{original_id}",
        json={"prompt": "How many weavers work with you?"},
        headers=_headers(world, "admin"),
    )
    assert reworded.status_code == 200, reworded.text
    body = reworded.json()

    assert body["action"] == "superseded", (
        "the prompt was written straight over the old one — every answer already recorded now reads "
        "as an answer to a question nobody was ever asked"
    )
    assert body["id"] != original_id, "the replacement must be a NEW row, not the answered one"
    assert body["prompt"] == "How many weavers work with you?"
    assert body["supersededQuestionId"] == original_id
    assert body["isActive"] is True
    assert body["sectionId"] == default_questionnaire["sectionId"]

    # THE ANSWERED QUESTION, UNTOUCHED, WITH ITS ORIGINAL WORDING AND A LINK TO ITS REPLACEMENT.
    retired = _global_question(client, world, original_id)
    assert retired is not None, "the original was deleted; its nineteen answers would be orphaned"
    assert retired["prompt"] == "How many looms do you own?"
    assert retired["isActive"] is False
    assert retired["retiredAt"] is not None
    assert retired["supersededById"] == body["id"], (
        "without the link this is indistinguishable from somebody deleting one question and adding "
        "another, which loses the fact that a correction happened at all"
    )

    # And the recorded answer still hangs off the wording it was given under — not moved onto the
    # replacement, which is the failure this whole rule exists to prevent.
    from prisma import Prisma

    own = Prisma()
    await own.connect()
    try:
        rows = await own.questionnaireresponse.find_many(
            where={"interviewId": default_questionnaire["interviewId"]}
        )
        by_question = {r.questionId: r.answerText for r in rows}
        assert by_question.get(original_id) == "12"
        assert body["id"] not in by_question, (
            "the answer followed the new wording; twelve looms just became twelve weavers"
        )
    finally:
        await own.disconnect()


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_an_unanswered_default_question_is_reworded_in_place_with_no_ceremony(
    client, world, default_questionnaire
):
    """The other half of the rule, and the reason it is not simply "questionnaires are frozen".

    A professor who typed a question wrongly, or spotted a translation error before the first
    interview, must be able to fix it. Superseding a question nobody has answered would litter the
    instrument with retired rows nobody ever saw and make the whole feature feel broken — which is
    how a protection gets worked around instead of used.
    """
    question_id = default_questionnaire["questions"]["Who taught you?"]

    reworded = client.patch(
        f"/api/questionnaire/questions/{question_id}",
        json={"prompt": "Who taught you this craft?"},
        headers=_headers(world, "admin"),
    )
    assert reworded.status_code == 200, reworded.text
    body = reworded.json()

    assert body["action"] == "updated"
    assert body["id"] == question_id, "an unanswered question must keep its identity"
    assert body["prompt"] == "Who taught you this craft?"
    assert body["isActive"] is True
    assert body["supersededById"] is None
    assert body["retiredAt"] is None


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_a_default_question_that_was_already_superseded_refuses_a_second_reword(
    client, world, default_questionnaire
):
    """``supersededById`` holds ONE id, so a chain cannot be written through the middle of itself.

    Rewording an already-superseded question would overwrite its existing link, and a reader would be
    left with a retired question pointing at a replacement two steps away with no way to know a step
    was missing. The professor's real intent is to edit the wording being asked TODAY, so the refusal
    says where that is rather than merely refusing.
    """
    original_id = default_questionnaire["questions"]["Where do you buy your yarn?"]

    # Supersede it here rather than leaning on the earlier test, so this test states its own whole
    # story and cannot pass for the wrong reason when it is run on its own with `-k`.
    first = client.patch(
        f"/api/questionnaire/questions/{original_id}",
        json={"prompt": "Where do you buy your cotton?"},
        headers=_headers(world, "admin"),
    )
    assert first.status_code == 200, first.text
    assert first.json()["action"] == "superseded"
    replacement_id = first.json()["id"]

    refused = client.patch(
        f"/api/questionnaire/questions/{original_id}",
        json={"prompt": "Where do you buy your silk?"},
        headers=_headers(world, "admin"),
    )
    assert refused.status_code == 409, refused.text
    assert "replaced by a newer one" in refused.json()["detail"]

    unchanged = _global_question(client, world, original_id)
    assert unchanged["prompt"] == "Where do you buy your yarn?"
    assert unchanged["supersededById"] == replacement_id, "the existing link survived the refusal"

    # The way forward the refusal names actually works: the CURRENT wording is editable, and since
    # nobody has answered the replacement it is edited in place rather than superseded again.
    onward = client.patch(
        f"/api/questionnaire/questions/{replacement_id}",
        json={"prompt": "Where do you buy your silk?"},
        headers=_headers(world, "admin"),
    )
    assert onward.status_code == 200, onward.text
    assert onward.json()["action"] == "updated"
    assert onward.json()["id"] == replacement_id


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_help_text_and_position_stay_editable_on_an_answered_default_question(
    client, world, default_questionnaire
):
    """WHAT THE RULE DOES NOT FREEZE, pinned so the fix cannot be widened into a wall.

    An answer is evidence and the WORDS it was given under are part of that evidence. Where the
    question sits in the section is not. Freezing position would mean a professor could never reorder
    a questionnaire once the first interview had been recorded, which for the global instrument means
    never — and the natural workaround (retire everything and retype it) destroys far more than the
    rule was protecting.
    """
    answered_id = default_questionnaire["questions"]["What does a kilo of yarn cost?"]
    before = _global_question(client, world, answered_id)
    assert before["isActive"] is True

    moved = client.patch(
        f"/api/questionnaire/questions/{answered_id}",
        json={"sortOrder": (before["sortOrder"] or 1) + 5},
        headers=_headers(world, "admin"),
    )
    assert moved.status_code == 200, moved.text
    assert moved.json()["action"] == "updated", "moving a question is not rewording it"
    assert moved.json()["id"] == answered_id, "a reposition must not fork the question"
    assert moved.json()["sortOrder"] == (before["sortOrder"] or 1) + 5
    assert moved.json()["prompt"] == "What does a kilo of yarn cost?"
    assert moved.json()["isActive"] is True
    assert moved.json()["supersededById"] is None


@pytest.mark.skipif(not _LOCAL, reason="needs a LOCAL database")
@pytest.mark.anyio
async def test_removing_an_answered_default_question_stamps_that_it_carried_evidence(
    client, world, default_questionnaire
):
    """DELETE here has always been a retire; what it never did was SAY which retirement this was.

    An answered question switched off and one an admin unticked the day it was typed came out of this
    endpoint identical, so afterwards there was no way to tell a row that carries evidence from a row
    that was a mistake. ``retiredAt`` records the difference — the same distinction
    ``QuestionnaireFormQuestion.retiredAt`` draws in the custom builder.
    """
    unanswered_id = default_questionnaire["questions"]["Who taught you?"]
    removed = client.delete(
        f"/api/questionnaire/questions/{unanswered_id}", headers=_headers(world, "admin")
    )
    assert removed.status_code == 204, removed.text
    gone = _global_question(client, world, unanswered_id)
    assert gone["isActive"] is False
    assert gone["retiredAt"] is None, (
        "nobody answered this one, so it was switched off rather than retired — a stamp here would "
        "make the column mean 'inactive', which the isActive column already means"
    )

    answered_id = default_questionnaire["questions"]["What does a kilo of yarn cost?"]
    retired = client.delete(
        f"/api/questionnaire/questions/{answered_id}", headers=_headers(world, "admin")
    )
    assert retired.status_code == 204, retired.text
    kept = _global_question(client, world, answered_id)
    assert kept is not None, "an answered question must never actually be deleted"
    assert kept["isActive"] is False
    assert kept["retiredAt"] is not None, (
        "this question carries a recorded answer, so it was retired rather than merely switched "
        "off, and the record has to say which"
    )
    assert kept["supersededById"] is None, "nothing replaced it; it was removed, not reworded"

    # And the supersede path stamps it too — a superseded question is retired BECAUSE it carried
    # evidence, which is the whole reason its wording was not overwritten.
    superseded_id = default_questionnaire["questions"]["How many looms do you own?"]
    already = _global_question(client, world, superseded_id)
    assert already["retiredAt"] is not None
