"""The .docx's NATIVE charts: the parts, the schema sequences, and the phone's copy of both.

A ``c:chart`` fails in a way no other part of this package does. A picture that is wrong is
visibly wrong — a red X, a stretched face, a missing figure. A chart that is wrong opens as a
correctly sized EMPTY RECTANGLE, or as a "we found a problem with some content" dialog offering
to repair the document by deleting it, and both look to the reader like the report simply does
not have that figure. So every check here is about a failure whose only symptom is absence:

* a part declared ``application/xml`` instead of the chart content type — Word ignores it,
* a relationship id shared with an image — Word drops both parts and says nothing,
* a shuffled schema sequence — the part is invalid and Word offers to repair,
* ``c:cat``/``c:val`` without their literal caches — Word renders from the cache and only opens
  the workbook when the reader asks to edit, so a chart without one plots nothing on every
  machine that has not opened it, which is every machine,
* a number written in exponent form — Word reads ``1.2E7`` in a ``c:v`` as TEXT and the point
  leaves the plot.

The last section reads ``DocxWriter.kt`` as source text, exactly as ``test_report_parity.py``
does and for the same reason: the phone must write the same document, nothing in a compiler says
so, and the drift that actually happens is a constant edited on one side only.
"""

import re
import zipfile
from io import BytesIO
from xml.etree import ElementTree as ET

import openpyxl
import pytest

from app.services import report_docx
from app.services.report_chart import clean_series
from app.services.report_docx import DocxWriter, chart_space_xml, render_docx
from app.services.report_model import (
    ChartBlock,
    ChartKind,
    DocumentBuilder,
    MapBlock,
    ReportMeta,
    ReportTheme,
)
from tests.test_report_parity import _kotlin

C = "{http://schemas.openxmlformats.org/drawingml/2006/chart}"
CT = "{http://schemas.openxmlformats.org/package/2006/content-types}"
REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
CHART_TYPE = "application/vnd.openxmlformats-officedocument.drawingml.chart+xml"
XLSX_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

#: One realistic series: a decimal, a round number, a negative (a credited cost line) and a value
#: large enough that a naive float-to-string would switch to exponent notation.
SERIES = (("Material", 562.5), ("Labour", 900.0), ("Transport", -40.0), ("Dye", 12345678.0))


def _meta() -> ReportMeta:
    return ReportMeta(title="Workshop Report", generated_at="2026-08-07T00:00:00Z")


def _render(*blocks) -> zipfile.ZipFile:
    builder = DocumentBuilder(meta=_meta())
    for block in blocks:
        builder.add(block)
    data, _dropped = render_docx(builder.build(), lambda _ref: None)
    return zipfile.ZipFile(BytesIO(data))


def _chart(kind: ChartKind, **kw) -> ChartBlock:
    base = {"series": SERIES, "title": f"{kind.value} figure", "unit": "INR"}
    base.update(kw)
    return ChartBlock(kind=kind, **base)


def _local(element) -> str:
    return element.tag.rsplit("}", 1)[-1]


def _assert_in_schema_order(element, sequence: list[str], where: str) -> None:
    """Every child of ``element`` must appear in ``sequence``, in that order.

    A subsequence check rather than an equality check, because every one of these schema
    sequences is mostly optional elements — what is being pinned is the ORDER, which is what Word
    validates and what a hand-written part gets wrong.
    """
    children = [_local(child) for child in element]
    for name in children:
        assert name in sequence, f"{where}: <{name}> is not a legal child ({sequence})"
    positions = [sequence.index(name) for name in children]
    assert positions == sorted(positions), (
        f"{where}: children are out of schema sequence.\n  got      {children}\n"
        f"  schema   {sequence}"
    )


# --------------------------------------------------------------------------------------
# The parts
# --------------------------------------------------------------------------------------


@pytest.mark.parametrize("kind", list(ChartKind))
def test_every_chart_kind_becomes_a_native_chart_part(kind):
    """All five kinds have a Word form, so none of them should still be a picture."""
    z = _render(_chart(kind))
    assert "word/charts/chart1.xml" in z.namelist()
    assert "word/charts/_rels/chart1.xml.rels" in z.namelist()
    assert "word/embeddings/chart1.xlsx" in z.namelist()
    assert not [n for n in z.namelist() if n.startswith("word/media/")], \
        f"{kind.value} still rasterised"


def test_every_chart_part_is_well_formed_xml():
    z = _render(*[_chart(kind) for kind in ChartKind])
    for name in z.namelist():
        if name.endswith((".xml", ".rels")):
            ET.fromstring(z.read(name))       # raises on malformed XML
    assert z.testzip() is None


def test_the_chart_part_carries_the_chart_content_type_and_not_generic_xml():
    """A chart part ends in .xml, so the Default already covers it — as application/xml, which
    Word ignores completely, leaving a correctly sized blank rectangle in the document."""
    z = _render(_chart(ChartKind.BAR), _chart(ChartKind.PIE))
    ct = ET.fromstring(z.read("[Content_Types].xml"))
    declared = {o.get("PartName"): o.get("ContentType") for o in ct.findall(CT + "Override")}
    assert declared["/word/charts/chart1.xml"] == CHART_TYPE
    assert declared["/word/charts/chart2.xml"] == CHART_TYPE
    defaults = {d.get("Extension"): d.get("ContentType") for d in ct.findall(CT + "Default")}
    assert defaults["xlsx"] == XLSX_TYPE


def test_a_document_with_no_chart_declares_no_chart_types():
    """The xlsx Default and the chart Overrides must not appear in a report that has none."""
    z = _render()
    ct = z.read("[Content_Types].xml").decode()
    assert "xlsx" not in ct and "chart" not in ct
    assert not [n for n in z.namelist() if "chart" in n or "embeddings" in n]


def test_chart_relationship_ids_cannot_collide_with_image_ids():
    """The failure this guards is the one the module docstring names: two relationships with one
    id, and Word drops BOTH parts without complaint. Images derive their id arithmetically from
    the media count, so a chart emitted between two photographs is exactly where a shared counter
    would hand out a duplicate."""
    from app.services.report_model import ImageBlock, ImageRef
    from tests.test_report_docx import _png

    photos = {"a": _png(40, 30), "b": _png(40, 30)}
    builder = DocumentBuilder(meta=_meta())
    builder.add(ImageBlock(image=ImageRef("a", 40, 30, mime_type="image/png")))
    builder.add(_chart(ChartKind.BAR))
    builder.add(ImageBlock(image=ImageRef("b", 40, 30, mime_type="image/png")))
    builder.add(_chart(ChartKind.PIE))
    data, dropped = render_docx(builder.build(), lambda ref: photos.get(ref.source))
    assert dropped == []
    z = zipfile.ZipFile(BytesIO(data))

    rels = ET.fromstring(z.read("word/_rels/document.xml.rels"))
    ids = [r.get("Id") for r in rels.findall(REL + "Relationship")]
    assert len(ids) == len(set(ids)), f"duplicate relationship id in {ids}"
    assert {"cId1", "cId2", "rId6", "rId7"} <= set(ids), (
        f"a chart emitted between two photographs disturbed the image numbering: {ids}"
    )


def test_every_chart_reference_resolves_to_a_part_that_exists():
    z = _render(*[_chart(kind) for kind in ChartKind])
    doc = z.read("word/document.xml").decode()
    rels = ET.fromstring(z.read("word/_rels/document.xml.rels"))
    declared = {r.get("Id"): r.get("Target") for r in rels.findall(REL + "Relationship")}
    used = set(re.findall(r'r:id="(cId\d+)"', doc))
    assert len(used) == 5
    for rid in used:
        assert rid in declared, f"{rid} used in document.xml but never declared"
        assert f"word/{declared[rid]}" in z.namelist()


def test_each_chart_part_points_at_its_own_workbook():
    """chart2 linked to chart1's workbook opens "Edit Data" on the wrong figure's numbers."""
    z = _render(_chart(ChartKind.BAR), _chart(ChartKind.LINE))
    for n in (1, 2):
        rels = ET.fromstring(z.read(f"word/charts/_rels/chart{n}.xml.rels"))
        rel = rels.find(REL + "Relationship")
        assert rel.get("Target") == f"../embeddings/chart{n}.xlsx"
        # The embedded-package type, not oleObject and not image. A wrong type still draws the
        # chart (it renders from its cache) and breaks Edit Data with a missing-file dialog.
        assert rel.get("Type").endswith("/package")
        assert rel.get("Id") in z.read(f"word/charts/chart{n}.xml").decode()


def test_the_map_stays_a_raster():
    """It is a pinned projection, not a Word chart type. Whatever happens to the charts, the map
    must keep going through the picture path or the locator vanishes from the report."""
    z = _render(MapBlock(title="Where the workshop was held"))
    assert not [n for n in z.namelist() if n.startswith("word/charts/")]
    # Either a media part (the assets are present) or nothing at all (they are not) — but never
    # a chart part.
    doc = z.read("word/document.xml").decode()
    assert "<c:chart " not in doc


# --------------------------------------------------------------------------------------
# The schema sequences — where a wrong order costs a repair dialog
# --------------------------------------------------------------------------------------

_CHART_SPACE_SEQ = ["date1904", "lang", "roundedCorners", "style", "clrMapOvr", "pivotSource",
                    "protection", "chart", "spPr", "txPr", "externalData", "printSettings",
                    "userShapes", "extLst"]
_CHART_SEQ = ["title", "autoTitleDeleted", "pivotFmts", "view3D", "floor", "sideWall", "backWall",
              "plotArea", "legend", "plotVisOnly", "dispBlanksAs", "showDLblsOverMax", "extLst"]
_DLBLS_SEQ = ["numFmt", "spPr", "txPr", "dLblPos", "showLegendKey", "showVal", "showCatName",
              "showSerName", "showPercent", "showBubbleSize", "separator", "showLeaderLines",
              "leaderLines"]
_BAR_SER_SEQ = ["idx", "order", "tx", "spPr", "invertIfNegative", "pictureOptions", "dPt",
                "dLbls", "trendline", "errBars", "cat", "val", "shape", "extLst"]
_LINE_SER_SEQ = ["idx", "order", "tx", "spPr", "marker", "dPt", "dLbls", "trendline", "errBars",
                 "cat", "val", "smooth", "extLst"]
_PIE_SER_SEQ = ["idx", "order", "tx", "spPr", "explosion", "dPt", "dLbls", "cat", "val", "extLst"]
_CAT_AX_SEQ = ["axId", "scaling", "delete", "axPos", "majorGridlines", "minorGridlines", "title",
               "numFmt", "majorTickMark", "minorTickMark", "tickLblPos", "spPr", "txPr", "crossAx",
               "crosses", "crossesAt", "auto", "lblAlgn", "lblOffset", "tickLblSkip",
               "tickMarkSkip", "noMultiLvlLbl", "extLst"]
_VAL_AX_SEQ = ["axId", "scaling", "delete", "axPos", "majorGridlines", "minorGridlines", "title",
               "numFmt", "majorTickMark", "minorTickMark", "tickLblPos", "spPr", "txPr", "crossAx",
               "crosses", "crossesAt", "crossBetween", "majorUnit", "minorUnit", "dispUnits",
               "extLst"]
_SER_SEQ_FOR = {
    "barChart": _BAR_SER_SEQ,
    "lineChart": _LINE_SER_SEQ,
    "pieChart": _PIE_SER_SEQ,
    "doughnutChart": _PIE_SER_SEQ,
}


@pytest.mark.parametrize("kind", list(ChartKind))
def test_the_chart_part_follows_every_schema_sequence(kind):
    root = ET.fromstring(chart_space_xml(_chart(kind), list(SERIES), ReportTheme())
                         .split("?>", 1)[1])
    _assert_in_schema_order(root, _CHART_SPACE_SEQ, "c:chartSpace")
    chart = root.find(C + "chart")
    _assert_in_schema_order(chart, _CHART_SEQ, "c:chart")

    group = next(child for child in chart.find(C + "plotArea")
                 if _local(child).endswith("Chart"))
    ser = group.find(C + "ser")
    _assert_in_schema_order(ser, _SER_SEQ_FOR[_local(group)], f"c:ser of c:{_local(group)}")
    for dlbls in root.iter(C + "dLbls"):
        _assert_in_schema_order(dlbls, _DLBLS_SEQ, "c:dLbls")
    for ax, seq in ((C + "catAx", _CAT_AX_SEQ), (C + "valAx", _VAL_AX_SEQ)):
        for element in root.iter(ax):
            _assert_in_schema_order(element, seq, ax)


@pytest.mark.parametrize("kind", list(ChartKind))
def test_openpyxls_own_chart_model_accepts_the_part(kind):
    """A second opinion from a library whose descriptors are generated from the ECMA schema.

    ``c:externalData`` is lifted out first: openpyxl cannot round-trip it (its ``id`` descriptor
    never registers the ``r:`` namespace, so reading any real Word chart hits the same wall), and
    that is a gap in the reader rather than in this part — which the assertion below pins.
    """
    from openpyxl.chart.chartspace import ChartSpace

    root = ET.fromstring(chart_space_xml(_chart(kind), list(SERIES), ReportTheme())
                         .split("?>", 1)[1])
    external = root.find(C + "externalData")
    assert external is not None and list(external.attrib.values()) == ["rId1"]
    root.remove(external)

    space = ChartSpace.from_tree(root)          # raises on an element it does not know
    group = space.chart.plotArea._charts[0]
    assert type(group).__name__ == {
        ChartKind.BAR: "BarChart", ChartKind.HORIZONTAL_BAR: "BarChart",
        ChartKind.PIE: "PieChart", ChartKind.DONUT: "DoughnutChart",
        ChartKind.LINE: "LineChart",
    }[kind]


def test_a_doughnut_never_carries_a_data_label_position():
    """CT_DoughnutChart has no legal value for dLblPos and Word refuses the part outright."""
    xml = chart_space_xml(_chart(ChartKind.DONUT), list(SERIES), ReportTheme())
    assert "<c:doughnutChart>" in xml
    assert "dLblPos" not in xml


def test_the_bar_kinds_choose_their_direction_and_the_horizontal_one_reads_downward():
    """Word's default bar chart runs bottom-up, which silently reverses a cost sheet."""
    vertical = chart_space_xml(_chart(ChartKind.BAR), list(SERIES), ReportTheme())
    horizontal = chart_space_xml(_chart(ChartKind.HORIZONTAL_BAR), list(SERIES), ReportTheme())
    assert '<c:barDir val="col"/>' in vertical
    assert '<c:orientation val="minMax"/>' in vertical.split("<c:catAx>")[1]
    assert '<c:barDir val="bar"/>' in horizontal
    assert '<c:orientation val="maxMin"/>' in horizontal.split("<c:catAx>")[1]


# --------------------------------------------------------------------------------------
# The data — the cache, the workbook, and the numbers in both
# --------------------------------------------------------------------------------------


@pytest.mark.parametrize("kind", list(ChartKind))
def test_the_literal_cache_holds_every_point(kind):
    """Word renders from the cache and opens the workbook only on "Edit Data". A chart with the
    references and no cache plots NOTHING on every machine that has not opened it."""
    block = _chart(kind)
    expected, _notes = clean_series(block)
    root = ET.fromstring(chart_space_xml(block, expected, ReportTheme()).split("?>", 1)[1])

    cats = root.find(f".//{C}cat/{C}strRef/{C}strCache")
    vals = root.find(f".//{C}val/{C}numRef/{C}numCache")
    assert cats.find(C + "ptCount").get("val") == str(len(expected))
    assert vals.find(C + "ptCount").get("val") == str(len(expected))
    assert [p.find(C + "v").text for p in cats.findall(C + "pt")] == [ln for ln, _v in expected]
    assert [float(p.find(C + "v").text) for p in vals.findall(C + "pt")] == \
        [v for _ln, v in expected]


@pytest.mark.parametrize("kind", list(ChartKind))
def test_the_native_chart_plots_exactly_what_the_rasteriser_would_have_drawn(kind):
    """One filter, in ``report_chart.clean_series``, for both renderers.

    If the chart writer filtered for itself, a NaN that one drops and the other keeps would put a
    category in the .docx's editable chart that is missing from the .pdf's picture of it — two
    documents of one workshop disagreeing about which cost heads exist.
    """
    block = ChartBlock(kind=kind, title="t", series=(
        ("Material", 562.5), ("Broken", float("nan")), ("Credit", -40.0), ("Labour", 900.0),
    ))
    kept, notes = clean_series(block)
    z = _render(block)
    root = ET.fromstring(z.read("word/charts/chart1.xml"))
    plotted = [p.find(C + "v").text
               for p in root.findall(f".//{C}cat/{C}strRef/{C}strCache/{C}pt")]
    assert plotted == [label for label, _v in kept]
    assert "Broken" not in plotted
    if kind.is_circular:
        assert "Credit" not in plotted, "a negative cannot be a slice"
    # And whatever was dropped is named in the document, because a native chart has no corner to
    # print it in the way the PNG does.
    doc = z.read("word/document.xml").decode()
    for note in notes:
        assert note.split(":")[0] in doc


def test_the_embedded_workbook_is_a_real_xlsx_that_opens():
    z = _render(_chart(ChartKind.BAR))
    book = openpyxl.load_workbook(BytesIO(z.read("word/embeddings/chart1.xlsx")))
    sheet = book["Sheet1"]
    assert [[c.value for c in row] for row in sheet.iter_rows()] == [
        [None, "BAR figure"],
        ["Material", 562.5], ["Labour", 900], ["Transport", -40], ["Dye", 12345678],
    ]


def test_the_workbook_ranges_point_at_the_cells_the_workbook_actually_has():
    """The formulas and the sheet layout are two halves of one decision. Changed apart, "Edit
    Data" opens a sheet whose ranges select empty cells."""
    z = _render(_chart(ChartKind.LINE))
    xml = z.read("word/charts/chart1.xml").decode()
    assert "<c:f>Sheet1!$A$2:$A$5</c:f>" in xml       # four categories -> rows 2..5
    assert "<c:f>Sheet1!$B$2:$B$5</c:f>" in xml
    assert "<c:f>Sheet1!$B$1</c:f>" in xml            # the series name
    sheet = openpyxl.load_workbook(BytesIO(z.read("word/embeddings/chart1.xlsx")))["Sheet1"]
    assert sheet["A2"].value == "Material" and sheet["B5"].value == 12345678
    assert sheet["B1"].value == "LINE figure"


def test_a_large_value_is_never_written_in_exponent_form():
    """Word reads ``1.2345678E7`` in a ``c:v`` as TEXT and the point silently leaves the plot.
    Python only switches notation at 1e16, Kotlin's ``Double.toString`` at 1e7 — so this is the
    single most likely way the two writers could disagree about what is in a chart."""
    assert report_docx._cell_number(12345678.0) == "12345678"
    assert report_docx._cell_number(1e15) == "1000000000000000"
    assert report_docx._cell_number(0.000001) == "0.000001"
    assert report_docx._cell_number(-40.0) == "-40"
    assert report_docx._cell_number(562.5) == "562.5"
    for text in (report_docx._cell_number(v) for v in (1e14, 1e15, 12345678.9, 1e-6)):
        assert "e" not in text.lower(), text


def test_an_indic_category_reaches_the_chart_that_the_rasteriser_would_have_dropped():
    """The one place the native chart is strictly better than the PNG it replaces.

    ``report_raster``'s five-by-seven face is ASCII only and drops what it cannot draw, so a cost
    head named in Odia was previously an EMPTY slot under a bar. A chart part carries the string
    and ``a:cs`` tells Word which face to shape it with.
    """
    block = ChartBlock(kind=ChartKind.BAR, title="Cost", series=(("ସମ୍ବଲପୁରୀ", 5.0), ("Dye", 3.0)))
    z = _render(block)
    xml = z.read("word/charts/chart1.xml").decode("utf-8")
    assert "ସମ୍ବଲପୁରୀ" in xml
    assert f'<a:cs typeface="{ReportTheme().complex_font}"/>' in xml
    sheet = openpyxl.load_workbook(BytesIO(z.read("word/embeddings/chart1.xlsx")))["Sheet1"]
    assert sheet["A2"].value == "ସମ୍ବଲପୁରୀ"


def test_xml_entities_in_a_category_are_escaped_in_both_the_chart_and_the_workbook():
    block = ChartBlock(kind=ChartKind.PIE, title='R&D "core"', series=(("A<B", 1.0), ("C&D", 2.0)))
    z = _render(block)
    ET.fromstring(z.read("word/charts/chart1.xml"))
    book = openpyxl.load_workbook(BytesIO(z.read("word/embeddings/chart1.xlsx")))
    assert [c.value for c in book["Sheet1"]["A"]][1:] == ["A<B", "C&D"]


def test_the_workbook_is_byte_stable_across_renders():
    """The .docx is a deliverable that gets re-generated, diffed and re-uploaded. A zip stamped
    with "now" makes a figure nobody edited show as changed."""
    first = _render(_chart(ChartKind.BAR)).read("word/embeddings/chart1.xlsx")
    second = _render(_chart(ChartKind.BAR)).read("word/embeddings/chart1.xlsx")
    assert first == second


# --------------------------------------------------------------------------------------
# Degrading, and the one input that must
# --------------------------------------------------------------------------------------


@pytest.mark.parametrize("kind", list(ChartKind))
def test_a_chart_with_no_usable_values_keeps_its_picture(kind):
    """``report_chart`` draws that case as a framed panel reading "No values recorded.", which is
    a statement ABOUT THE RECORD. A native chart with no points draws an empty plot area and says
    nothing at all, so here the raster is the better document — degrading is the correct answer,
    and dropping the figure never is."""
    builder = DocumentBuilder(meta=_meta())
    builder.add(ChartBlock(kind=kind, series=(), title="Nothing recorded"))
    writer = DocxWriter(builder.build(), lambda _ref: None)
    z = zipfile.ZipFile(BytesIO(writer.build()))

    assert writer.rasterised_charts == ["Nothing recorded"]
    assert not [n for n in z.namelist() if n.startswith("word/charts/")]
    assert [n for n in z.namelist() if n.startswith("word/media/")] == ["word/media/image1.png"]
    assert "<w:drawing>" in z.read("word/document.xml").decode()


def test_the_figure_title_and_caption_are_real_text_on_both_paths():
    """Not drawn into the chart and not drawn into the bitmap. That is what lets a caption carry
    Odia, what makes it searchable, and what keeps a native chart sitting on the page exactly
    where the rasterised map above it does."""
    native = _render(_chart(ChartKind.BAR, title="Cost by head", caption="From the sheets."))
    doc = native.read("word/document.xml").decode()
    assert "Cost by head" in doc and "From the sheets." in doc
    # And Word must not invent a second title inside the frame from the series name.
    assert '<c:autoTitleDeleted val="1"/>' in native.read("word/charts/chart1.xml").decode()

    empty = _render(ChartBlock(kind=ChartKind.BAR, series=(), title="Cost by head",
                               caption="From the sheets."))
    doc = empty.read("word/document.xml").decode()
    assert "Cost by head" in doc and "From the sheets." in doc


def test_a_horizontal_bar_chart_is_given_room_for_every_category():
    """Word's response to axis labels that do not fit is to DROP them, silently and from the
    middle. A cost chart missing "Transport" between "Material" and "Labour" is not a smaller
    figure, it is a wrong one, and nothing in the document says so."""
    rows = 8
    z = _render(ChartBlock(
        kind=ChartKind.HORIZONTAL_BAR, title="Cost by head", unit="INR",
        series=tuple((f"Head {i}", float(i + 1)) for i in range(rows)),
    ))
    doc = z.read("word/document.xml").decode()
    cy = int(re.search(r'<wp:extent cx="\d+" cy="(\d+)"/>', doc).group(1))
    height_mm = cy / 36000
    assert height_mm >= report_docx._NATIVE_CHROME_MM + report_docx._NATIVE_ROW_MM * rows - 0.5
    # …and never taller than the box the picture path would have been allowed.
    assert height_mm <= (297 - 50) * 0.58 + 0.5


def test_a_chart_never_exceeds_the_text_column():
    z = _render(_chart(ChartKind.BAR, width_pct=100.0))
    doc = z.read("word/document.xml").decode()
    cx = int(re.search(r'<wp:extent cx="(\d+)"', doc).group(1))
    assert cx <= (160 * 36000) + 1        # A4 minus two 25 mm margins, at 36000 EMU/mm


# --------------------------------------------------------------------------------------
# The phone's copy — read as source text, exactly as test_report_parity.py does
# --------------------------------------------------------------------------------------


#: Kotlin string constants the chart XML is interpolated from, expanded so a literal comparison
#: is about the XML rather than about which side happened to name a namespace.
#:
#: The same idea as ``_kotlin``'s own stripping of escaped quotes and concatenated literals: what
#: is being compared is the bytes the two writers emit, and `"<c:chartSpace xmlns:c=\"$NS_C\""` is
#: the same bytes as the Python's f-string. Only true constants are expanded — anything derived
#: from the theme or the data stays interpolated and is excluded by ``_INTERPOLATED`` below.
_KOTLIN_CONSTANTS = {
    "$NS_C": report_docx._NS_C,
    "$NS_A": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "$NS_R": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "$NS_SS": report_docx._NS_SS,
    "$NS_CT": report_docx._NS_CT,
    "$NS_PKG_REL": report_docx._NS_PKG_REL,
    "$WORKBOOK_RID": report_docx._WORKBOOK_RID,
    "$CAT_AX_ID": str(report_docx._CAT_AX_ID),
    "$VAL_AX_ID": str(report_docx._VAL_AX_ID),
    "$CHART_PART_TYPE": report_docx._CHART_PART_TYPE,
    "$XLSX_MIME": report_docx._XLSX_MIME,
}


@pytest.fixture(scope="module")
def docx_kt() -> str:
    text = _kotlin("DocxWriter.kt")
    # Longest first, so "$NS_PKG_REL" is not eaten by a prefix of "$NS_P…".
    for name in sorted(_KOTLIN_CONSTANTS, key=len, reverse=True):
        text = text.replace(name, _KOTLIN_CONSTANTS[name])
    return text


@pytest.fixture(scope="module")
def chart_kt() -> str:
    return _kotlin("ReportChart.kt")


def test_the_phone_writes_native_charts_at_all(docx_kt):
    """The failure this file exists to prevent, stated plainly: a server that writes a chart and a
    phone that writes a picture hand a designer in a field a different document from the one the
    office downloads for the same workshop, and neither says so."""
    for symbol in ("chartSpaceXml", "chartWorkbookXlsx", "chartPartRels", "emitNativeChart",
                   "chartDrawing"):
        assert symbol in docx_kt, f"{symbol} is in report_docx.py but not in DocxWriter.kt"


def test_the_chart_constants_match(docx_kt):
    """A constant edited on one side is the drift that actually happens — see the module docstring
    of test_report_parity.py, which this section follows."""
    assert f'"{report_docx._CHART_RID_PREFIX}"' in docx_kt, "the chart rId prefix differs"
    assert str(report_docx._CAT_AX_ID) in docx_kt and str(report_docx._VAL_AX_ID) in docx_kt
    assert f'"{report_docx._WORKBOOK_RID}"' in docx_kt
    assert report_docx._CHART_PART_TYPE in docx_kt
    assert report_docx._XLSX_MIME in docx_kt
    assert str(report_docx._NATIVE_ROW_MM) in docx_kt
    assert str(report_docx._NATIVE_CHROME_MM) in docx_kt
    # The Kotlin escapes the backslashes of the Excel format code; _kotlin() strips the escaping
    # of quotes only, so compare on the digits that carry the meaning.
    assert "[>=10000000]" in docx_kt and "[>=100000]" in docx_kt, \
        "the Indian number format differs; a chart axis would group Western beside a cost table " \
        "that groups Indian"


#: Tags whose attribute values come from the theme, the data or the series length, and so cannot
#: appear as a literal in either source. Everything else in a chart part is a constant, and a
#: constant that differs between the two writers is a chart that differs between the two files.
_INTERPOLATED = {
    "c:f", "c:v", "c:pt", "c:ptCount", "c:idx", "c:axId", "c:crossAx", "c:numFmt",
    "a:srgbClr", "a:latin", "a:cs", "a:defRPr", "a:t", "c:orientation", "c:axPos", "c:barDir",
    "c:dLblPos", "c:crossBetween", "a:bodyPr", "c:showVal", "c:showPercent",
    # The Kotlin writes `<c:$tag>` from one branch shared by the pie and the doughnut, so the
    # literal element name is not in the source to find. test_the_circular_chart_tag_names_match
    # covers those two by name instead.
    "c:pieChart", "c:doughnutChart",
}


@pytest.mark.parametrize("kind", list(ChartKind))
def test_every_constant_tag_of_a_chart_part_appears_in_the_kotlin(kind, docx_kt):
    """Derived from the XML the Python actually emits rather than from a hand-kept list.

    A hand-kept list is a list somebody forgets to add to. Taking the tags out of the produced
    part means a new element added on the server is checked on the phone the moment it is written,
    which is the only time the fix is cheap.
    """
    xml = chart_space_xml(_chart(kind), list(SERIES), ReportTheme())
    tags = {t for t in re.findall(r"<[ca]:[A-Za-z0-9]+(?:[^<>]*?)/?>", xml)
            if re.match(r"<([ca]:[A-Za-z0-9]+)", t).group(1) not in _INTERPOLATED}
    # Prove the sweep is live rather than vacuously true: a chart part is forty-odd constant tags
    # and an _INTERPOLATED entry added carelessly could quietly empty this set.
    assert len(tags) >= 40, f"only {len(tags)} tags checked; _INTERPOLATED has grown too broad"
    missing = sorted(t for t in tags if t not in docx_kt)
    assert not missing, f"{kind.value}: emitted by report_docx.py, absent from DocxWriter.kt:\n" \
                        + "\n".join(missing)


def test_the_circular_chart_tag_names_match(docx_kt):
    """Checked apart from the tag sweep above because the Kotlin builds `<c:$tag>` from a local —
    a pie and a doughnut differ only in the element name and the hole, so one shared branch is
    correct there and the literal simply does not exist to compare."""
    assert '"doughnutChart"' in docx_kt and '"pieChart"' in docx_kt
    assert "<c:$tag>" in docx_kt and "</c:$tag>" in docx_kt, \
        "DocxWriter.kt no longer builds the circular chart tag from one branch; the sweep in " \
        "test_every_constant_tag_of_a_chart_part_appears_in_the_kotlin should now cover it"


def test_every_workbook_part_appears_in_the_kotlin(docx_kt):
    """The embedded workbook is hand-written on both sides precisely so it can be — openpyxl on
    the server and a hand-rolled part on the phone would be two different files inside two
    documents claiming to be the same report."""
    for part in ("[Content_Types].xml", "_rels/.rels", "xl/workbook.xml",
                 "xl/_rels/workbook.xml.rels", "xl/styles.xml", "xl/worksheets/sheet1.xml"):
        assert part in docx_kt, f"the phone's workbook is missing {part}"
    for token in ('t="inlineStr"', 'name="Sheet1"', 'builtinId="0"', 'patternType="gray125"',
                  "spreadsheetml.sheet.main+xml", "spreadsheetml.worksheet+xml"):
        assert token in docx_kt, f"the phone's workbook differs at {token}"


def test_both_writers_take_the_chart_geometry_from_the_rasteriser(chart_kt):
    """The native chart has no bitmap and so no intrinsic size. If the two sides derived its frame
    independently the .docx's figure would be a different SHAPE from the .pdf's picture of it."""
    assert "chartPixelBox" in chart_kt
    assert "internal fun cleanSeries" in chart_kt, \
        "cleanSeries must be visible to DocxWriter.kt or the phone filters its own series"
    assert "internal fun sliceColours" in chart_kt


def test_both_writers_fall_back_to_the_picture_for_an_empty_series(docx_kt):
    assert "rasterisedCharts" in docx_kt
    assert "renderChartPng" in docx_kt, "the phone must keep the raster path for the fallback"


def test_the_phone_never_writes_a_number_in_exponent_form(docx_kt):
    """Kotlin's ``Double.toString`` switches to exponent notation at 1e7 — a threshold every
    rupee figure in a cost chart crosses. This is the one place the port cannot be literal."""
    assert "cellNumber" in docx_kt
    assert "BigDecimal" in docx_kt, \
        "DocxWriter.kt must format chart values without Double.toString's exponent notation"
